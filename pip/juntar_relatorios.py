#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Junta os relatórios ACADÊMICO + PROFISSIONAL da UnB num ÚNICO xlsx, com uma
coluna 'Modalidade' nas abas (filtrável). Reusa a classificação/estilos de
investiga_quedas.py.

Saída: saida/relatorio_quedas_unb_completo.xlsx
"""
import os
from statistics import mean, median
from collections import Counter
import openpyxl
from openpyxl.styles import Font, Border, Side

import investiga_quedas as IQ
from investiga_quedas import (carregar, montar, ANOS, BASE, REC, LIM_RED, LIM_YEL,
                              FILLS, H, HFILL, BORDA, CEN, LFT, THIN, _hdr, _fmt_variacao)

SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'saida')
MOD_FILL = {'Acadêmico': openpyxl.styles.PatternFill('solid', fgColor='EAF2F8'),
            'Profissional': openpyxl.styles.PatternFill('solid', fgColor='FEF9E7')}


def build():
    art, names, roster, pa_artpe, pa_autor, meta = carregar()
    progs = []
    for modal, rotulo in (('ACAD', 'Acadêmico'), ('PROF', 'Profissional')):
        for p in montar(art, names, roster, pa_artpe, pa_autor, meta, modal=modal):
            p['modalidade'] = rotulo
            progs.append(p)
    return progs


def _row_style(ws, rr, ncols, modal, left_cols=()):
    for i in range(1, ncols + 1):
        c = ws.cell(rr, i); c.border = BORDA
        c.alignment = LFT if i in left_cols else CEN
    ws.cell(rr, 1).fill = MOD_FILL[modal]


# ------------------------------------------------------------------ abas
def aba_resumo(wb, progs):
    ws = wb.create_sheet('Resumo')
    ws.cell(1, 1, 'QUEDAS DE PRODUÇÃO — UnB · ACADÊMICOS + PROFISSIONAIS (dados brutos CAPES)').font = \
        Font(bold=True, size=14, color='2C3E50')
    ws.merge_cells('A1:H1')
    ws.cell(2, 1, 'Sinal: nº de artigos distintos por programa/ano (prod_intel_artpe). '
                  'Classe pela retenção limpa (artigos 2021-24 / 2017-20).').font = \
        Font(size=10, italic=True, color='555555')
    ws.merge_cells('A2:H2')
    _hdr(ws, ['Modalidade', 'Programas', '🟢 Verde', '🟡 Amarelo', '🔴 Vermelho',
              '⬛ Estrutural', 'Retenção mediana'], [14, 11, 9, 10, 11, 12, 15], row=4)
    r = 5
    for modal in ('Acadêmico', 'Profissional'):
        sub = [p for p in progs if p['modalidade'] == modal]
        c = Counter(p['classe'] for p in sub)
        rets = [p['ret'] for p in sub if p['ret'] is not None and p['classe'] != 'ESTRUTURAL']
        ws.append([modal, len(sub), c.get('VERDE', 0), c.get('AMARELO', 0),
                   c.get('VERMELHO', 0), c.get('ESTRUTURAL', 0),
                   round(median(rets), 2) if rets else '—'])
        _row_style(ws, r, 7, modal, left_cols=(1,)); r += 1
    tot = Counter(p['classe'] for p in progs)
    ws.append(['TOTAL', len(progs), tot.get('VERDE', 0), tot.get('AMARELO', 0),
               tot.get('VERMELHO', 0), tot.get('ESTRUTURAL', 0), ''])
    for i in range(1, 8):
        ws.cell(r, i).border = BORDA; ws.cell(r, i).font = Font(bold=True)
        ws.cell(r, i).alignment = LFT if i == 1 else CEN
    return ws


def aba_serie(wb, progs):
    ws = wb.create_sheet('Série anual (artigos)')
    headers = ['Modalidade', 'Programa', 'Área CAPES'] + [str(y) for y in ANOS] + \
              ['Nota', 'ret', '% variação', 'Classe', 'Observação']
    larg = [12, 30, 22] + [5] * len(ANOS) + [5, 6, 9, 11, 30]
    _hdr(ws, headers, larg)
    col_var = 3 + len(ANOS) + 3
    for p in progs:
        ret_disp = '—' if (p['ret'] is None or p.get('base_incompleta')) else round(p['ret'], 2)
        ws.append([p['modalidade'], p['programa'], p['area']] + [p['serie'][y] for y in ANOS] +
                  [p['nota'], ret_disp, (p['var_pct'] if p['var_pct'] is not None else '—'),
                   p['classe'], p['motivo']])
        rr = ws.max_row
        _row_style(ws, rr, len(headers), p['modalidade'], left_cols=(1, 2, 3, len(headers)))
        _fmt_variacao(ws.cell(rr, col_var), p['var_pct'])
        for y in REC:
            ws.cell(rr, 3 + ANOS.index(y) + 1).fill = FILLS[p['classe']]
        ws.cell(rr, len(headers) - 1).fill = FILLS[p['classe']]
    return ws


def aba_por_pesquisador(wb, progs):
    ws = wb.create_sheet('Produção por pesquisador')
    ws.cell(1, 1, 'Produção por pesquisador por ANO — (a) art/docente = artigos/n_doc; '
                  '(b) art/permanente = artigos c/ autor permanente / n_perm. '
                  '2022-24: n_doc/n_perm = roster 2021.').font = Font(italic=True, size=9, color='555555')
    ws.merge_cells('A1:Z1')
    headers = ['Modalidade', 'Programa', 'Métrica'] + [str(y) for y in ANOS]
    larg = [12, 30, 16] + [5] * len(ANOS)
    _hdr(ws, headers, larg, row=2)
    sep = Side(style='medium', color='888888')
    for p in progs:
        if p['classe'] == 'ESTRUTURAL':
            continue
        ws.append([p['modalidade'], p['programa'], 'art/docente'] + [
            (round(p['serie'][y] / p['ndoc'][y], 2) if p['ndoc'].get(y) else None) for y in ANOS])
        r1 = ws.max_row
        ws.append([p['modalidade'], p['programa'], 'art/permanente'] + [
            (round(p['artperm'][y] / p['nperm'][y], 2)
             if (p['artperm'].get(y) is not None and p['nperm'].get(y)) else None) for y in ANOS])
        r2 = ws.max_row
        for rr in (r1, r2):
            _row_style(ws, rr, len(headers), p['modalidade'], left_cols=(1, 2, 3))
            for y in REC:
                ws.cell(rr, 3 + ANOS.index(y) + 1).fill = FILLS[p['classe']]
        for i in range(1, len(headers) + 1):
            cc = ws.cell(r1, i)
            cc.border = Border(left=THIN, right=THIN, bottom=THIN, top=sep)
        ws.cell(r1, 1).fill = MOD_FILL[p['modalidade']]
    return ws


def aba_lista_vermelha(wb, progs):
    ws = wb.create_sheet('Lista vermelha')
    sub = [p for p in progs if p['classe'] in ('VERMELHO', 'ESTRUTURAL')]
    _hdr(ws, ['Modalidade', 'Programa', 'Área CAPES', 'Nota', 'Classe', 'Série 2017-20',
              'Série 2021-24', '% variação', 'Diagnóstico', 'Código'],
         [12, 28, 22, 5, 11, 16, 16, 9, 32, 15])
    for p in sub:
        s = p['serie']
        ws.append([p['modalidade'], p['programa'], p['area'], p['nota'], p['classe'],
                   '/'.join(str(s[y]) for y in BASE), '/'.join(str(s[y]) for y in REC),
                   (p['var_pct'] if p['var_pct'] is not None else '—'), p['motivo'], p['cd']])
        rr = ws.max_row
        _row_style(ws, rr, 10, p['modalidade'], left_cols=(1, 2, 3, 6, 7, 9))
        ws.cell(rr, 5).fill = FILLS[p['classe']]
        _fmt_variacao(ws.cell(rr, 8), p['var_pct'])
    return ws


def main():
    progs = build()
    # ordena: modalidade, classe, retenção
    ordem = {'VERMELHO': 0, 'ESTRUTURAL': 1, 'AMARELO': 2, 'VERDE': 3}
    progs.sort(key=lambda p: (p['modalidade'], ordem[p['classe']],
                              p['ret'] if p['ret'] is not None else 9))
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    aba_resumo(wb, progs)
    aba_serie(wb, progs)
    aba_por_pesquisador(wb, progs)
    aba_lista_vermelha(wb, progs)
    from openpyxl.worksheet.properties import PageSetupProperties
    for ws in wb.worksheets:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins.left = ws.page_margins.right = 0.2
    out = os.path.join(SAIDA, 'relatorio_quedas_unb_completo.xlsx')
    wb.save(out)
    c = Counter((p['modalidade'], p['classe']) for p in progs)
    print('Combinado salvo:', out)
    for modal in ('Acadêmico', 'Profissional'):
        print(f"  {modal}: " + ' '.join(f"{k}={c.get((modal,k),0)}"
              for k in ('VERDE', 'AMARELO', 'VERMELHO', 'ESTRUTURAL')))


if __name__ == '__main__':
    main()
