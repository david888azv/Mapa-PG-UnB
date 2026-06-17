#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Versão NACIONAL (todas as IES) da investigação de quedas de produção.
Mesma metodologia de investiga_quedas.py (sinal limpo = artigos distintos por
programa/ano via prod_intel_artpe), aplicada a TODOS os programas acadêmicos.

Pré-requisito:  python3 build_cache_nacional.py   (gera cache_nac/)
Metadados (IES, área, modalidade, situação) vêm de programas_*.csv (brutos).

Saídas (saida/):
  nacional_quedas.xlsx   — abas: Distribuição, Programas (todos), Resumo por IES,
                           Lista vermelha nacional, Metodologia
"""
import os, glob, json, time, collections
from statistics import mean, median
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

from investiga_quedas import (ANOS, BASE, REC, LIM_RED, LIM_YEL,
                              FILLS, COR, H, HFILL, BORDA, CEN, LFT,
                              _hdr, _fmt_variacao, is_acad, manter_modal, MODAL, MLABEL, SUF)

AQUI = os.path.dirname(os.path.abspath(__file__))
D = os.path.normpath(os.path.join(AQUI, '..', '..', 'dados_capes')) + os.sep
CACHE = os.path.join(AQUI, 'cache_nac')
SAIDA = os.path.join(AQUI, 'saida')


def carregar_meta_programas():
    """cd -> {sigla, uf, area, grande_area, modal, situacao} de programas_*.csv."""
    meta = {}
    cols = {'CD_PROGRAMA_IES', 'AN_BASE', 'SG_ENTIDADE_ENSINO', 'SG_UF_PROGRAMA',
            'NM_AREA_AVALIACAO', 'NM_GRANDE_AREA_CONHECIMENTO',
            'NM_MODALIDADE_PROGRAMA', 'DS_SITUACAO_PROGRAMA', 'NM_PROGRAMA_IES'}
    for fp in sorted(glob.glob(D + 'programas_2013a2016_*.csv') +
                     glob.glob(D + 'programas_2017a2020_*.csv')):
        df = pd.read_csv(fp, sep=';', encoding='latin-1', low_memory=False,
                         usecols=lambda c: c.strip().upper() in cols)
        df.columns = [c.strip().upper() for c in df.columns]
        for _, r in df.iterrows():
            cd = r['CD_PROGRAMA_IES']; an = int(r['AN_BASE'])
            m = meta.get(cd)
            if m is None:
                meta[cd] = {'sigla': r.get('SG_ENTIDADE_ENSINO', '?'),
                            'uf': r.get('SG_UF_PROGRAMA', '?'),
                            'area': str(r.get('NM_AREA_AVALIACAO', '')),
                            'grande': str(r.get('NM_GRANDE_AREA_CONHECIMENTO', '')),
                            'modal': str(r.get('NM_MODALIDADE_PROGRAMA', '')),
                            'situacao': str(r.get('DS_SITUACAO_PROGRAMA', '')),
                            'programa': str(r.get('NM_PROGRAMA_IES', '')),
                            '_an': an}
            elif an >= m['_an']:   # situação/nome do registro mais recente
                m['situacao'] = str(r.get('DS_SITUACAO_PROGRAMA', '')); m['_an'] = an
    return meta


def montar(art, names, ies, roster, meta):
    # detecção de sucessores por (sigla, programa)
    porg = collections.defaultdict(list)
    for cd, m in meta.items():
        if manter_modal(m['modal']):
            porg[(m['sigla'], m['programa'])].append(cd)
    sucesso_de = {}
    for _, cds in porg.items():
        if len(cds) < 2: continue
        desat = [c for c in cds if 'DESATIV' in (meta[c]['situacao'] or '').upper()]
        ativ  = [c for c in cds if 'DESATIV' not in (meta[c]['situacao'] or '').upper()]
        if desat and ativ:
            for a in ativ: sucesso_de[a] = desat[0]

    def ser(cd): return {y: art.get(f'{cd}|{y}', 0) for y in ANOS}
    progs = []
    for cd in sorted(set(c.split('|')[0] for c in art)):
        m = meta.get(cd)
        if not m or not manter_modal(m['modal']):
            continue
        s = ser(cd)
        base = mean(s[y] for y in BASE); rec = mean(s[y] for y in REC)
        ret = (rec / base) if base > 0 else None
        nota = roster.get(f'{cd}|2021', {}).get('nota') or roster.get(f'{cd}|2020', {}).get('nota')

        sit = (m['situacao'] or '').upper()
        # encerrou antes do período: sem produção em 2020 NEM em 2021-24 (última atividade ≤2019)
        encerrou_antes = (base > 0 and rec == 0 and s[2020] == 0)
        if 'DESATIV' in sit or cd in sucesso_de.values():
            classe = 'ESTRUTURAL'; motivo = 'em desativação / sucedido por novo código'
        elif encerrou_antes:
            classe = 'ESTRUTURAL'; motivo = 'encerrado/recodificado (sem produção desde 2020)'
        elif ret is None:
            classe = 'ESTRUTURAL'; motivo = 'sem base 2017-20 (curso novo)'
        elif ret < LIM_RED:
            classe = 'VERMELHO'
            rec_tail = mean(s[y] for y in (2022, 2023, 2024))
            motivo = ('subnotificação (reportou 2021; 2022-24 ~zero)'
                      if base > 0 and s[2021] >= 0.5 * base and rec_tail < 0.25 * base
                      else 'queda acentuada')
        elif ret < LIM_YEL:
            classe = 'AMARELO'; motivo = 'queda moderada'
        else:
            classe = 'VERDE'; motivo = 'manteve ou cresceu'

        antigo = sucesso_de.get(cd)
        base_incompleta = bool(antigo) and sum(1 for y in BASE if s[y] == 0) >= 2
        var_pct = (round((ret - 1) * 100, 1) if ret is not None else None)
        if base_incompleta:
            var_pct = None; motivo = f'novo curso (sucessor) — base 2017-20 incompleta'

        progs.append({'cd': cd, 'sigla': m['sigla'], 'uf': m['uf'],
                      'programa': names.get(cd, m['programa']), 'area': m['area'],
                      'nota': nota, 'serie': s, 'base': base, 'rec': rec, 'ret': ret,
                      'var_pct': var_pct, 'classe': classe, 'motivo': motivo,
                      'base_incompleta': base_incompleta})
    ordem = {'VERMELHO': 0, 'ESTRUTURAL': 1, 'AMARELO': 2, 'VERDE': 3}
    progs.sort(key=lambda p: (ordem[p['classe']], p['ret'] if p['ret'] is not None else 9))
    return progs


# --------------------------------------------------------------------- Excel
def aba_distribuicao(wb, progs, rets, nac):
    ws = wb.create_sheet('Distribuição')
    from collections import Counter
    cont = Counter(p['classe'] for p in progs)
    ies_n = len(set(p['sigla'] for p in progs))
    n = len(rets)
    linhas = [
        (f'INVESTIGAÇÃO DE QUEDAS — NACIONAL · programas {MLABEL.upper()} (todas as IES)', 14, True, '2C3E50'),
        ('Sinal: nº de artigos distintos por programa por ano (prod_intel_artpe, dados brutos 2013-2024).', 10, False, '555555'),
        ('', 9, False, '000000'),
        (f'Programas {MLABEL}: {len(progs)}  |  IES distintas: {ies_n}', 11, True, '000000'),
        (f"   🟢 VERDE (ret≥{LIM_YEL}): {cont.get('VERDE',0)}", 11, False, '1E8449'),
        (f"   🟡 AMARELO ({LIM_RED}–{LIM_YEL}): {cont.get('AMARELO',0)}", 11, False, '9A7D0A'),
        (f"   🔴 VERMELHO (<{LIM_RED}): {cont.get('VERMELHO',0)}", 11, True, 'C0392B'),
        (f"   ⬛ ESTRUTURAL: {cont.get('ESTRUTURAL',0)}", 11, False, '566573'),
        ('', 9, False, '000000'),
        (f'Distribuição da retenção limpa (rec 2021-24 / base 2017-20), n={n} {MLABEL} não-desativados:', 10, True, '2C3E50'),
        (f'   p5={np.percentile(rets,5):.2f}  p10={np.percentile(rets,10):.2f}  p25={np.percentile(rets,25):.2f}  '
         f'mediana={median(rets):.2f}  p75={np.percentile(rets,75):.2f}  p90={np.percentile(rets,90):.2f}', 10, False, '333333'),
        ('', 9, False, '000000'),
        (f'ACHADO: retenção limpa mediana nacional dos programas {MLABEL} ≈ {median(rets):.2f}. A "queda"', 11, True, 'C0392B'),
        ('da camada pré-processada/mediana nacional (de ~8 p/ ~3 art/doc) é artefato de atribuição ao', 10, False, '555555'),
        ('roster de 2021, não perda real. Ver investiga_quedas.py / aba Metodologia do relatório UnB.', 10, False, '555555'),
        ('', 9, False, '000000'),
        ('Trajetória NACIONAL (mediana art/doc/ano, camada pré-processada — mostra o artefato):', 10, True, '2C3E50'),
    ]
    r = 1
    for txt, sz, b, c in linhas:
        ws.cell(r, 1, txt).font = Font(size=sz, bold=b, color=c)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        r += 1
    ws.cell(r, 1, 'Ano').font = H; ws.cell(r, 1).fill = HFILL
    for j, y in enumerate(ANOS, 2):
        c = ws.cell(r, j, y); c.font = H; c.fill = HFILL; c.alignment = CEN
    ws.cell(r+1, 1, 'Mediana nac.').font = Font(bold=True)
    for j, y in enumerate(ANOS, 2):
        ws.cell(r+1, j, nac.get(y)).alignment = CEN
    ws.column_dimensions['A'].width = 16
    return ws


def aba_programas(wb, progs):
    ws = wb.create_sheet('Programas')
    headers = ['IES', 'UF', 'Programa', 'Área CAPES'] + [str(y) for y in ANOS] + \
              ['Nota', 'ret', '% variação', 'Classe']
    larg = [9, 4, 30, 22] + [4]*len(ANOS) + [4, 5, 9, 11]
    _hdr(ws, headers, larg)
    col_var = 4 + len(ANOS) + 3
    for p in progs:
        ret_disp = '—' if (p['ret'] is None or p['base_incompleta']) else round(p['ret'], 2)
        ws.append([p['sigla'], p['uf'], p['programa'], p['area']] + [p['serie'][y] for y in ANOS] +
                  [p['nota'], ret_disp, (p['var_pct'] if p['var_pct'] is not None else '—'), p['classe']])
        rr = ws.max_row
        for i in range(1, len(headers)+1):
            c = ws.cell(rr, i); c.border = BORDA
            c.alignment = LFT if i in (3, 4) else CEN
        _fmt_variacao(ws.cell(rr, col_var), p['var_pct'])
        for y in REC:
            ws.cell(rr, 4 + ANOS.index(y) + 1).fill = FILLS[p['classe']]
        ws.cell(rr, len(headers)).fill = FILLS[p['classe']]
    return ws


def aba_resumo_ies(wb, progs):
    ws = wb.create_sheet('Resumo por IES')
    agg = collections.defaultdict(lambda: collections.Counter())
    uf = {}
    for p in progs:
        agg[p['sigla']][p['classe']] += 1; uf[p['sigla']] = p['uf']
    _hdr(ws, ['IES', 'UF', 'Programas', '🟢 Verde', '🟡 Amarelo', '🔴 Vermelho', '⬛ Estrutural',
              '% verde'], [12, 4, 10, 9, 10, 11, 11, 8])
    rows = []
    for sg, c in agg.items():
        tot = sum(c.values())
        rows.append((sg, uf[sg], tot, c['VERDE'], c['AMARELO'], c['VERMELHO'], c['ESTRUTURAL'],
                     round(100*c['VERDE']/tot, 0) if tot else 0))
    rows.sort(key=lambda r: (-r[2], r[0]))   # por nº de programas desc
    for row in rows:
        ws.append(list(row))
        rr = ws.max_row
        for i in range(1, 9):
            ws.cell(rr, i).border = BORDA; ws.cell(rr, i).alignment = CEN if i != 1 else LFT
        if row[5]:  # tem vermelho
            ws.cell(rr, 6).fill = FILLS['VERMELHO']
    return ws


def aba_lista_vermelha(wb, progs):
    ws = wb.create_sheet('Lista vermelha nacional')
    reds = [p for p in progs if p['classe'] == 'VERMELHO']
    _hdr(ws, ['IES', 'UF', 'Programa', 'Área CAPES', 'Nota', 'Série 2017-20', 'Série 2021-24',
              '% variação', 'Diagnóstico', 'Código'], [10, 4, 28, 22, 5, 16, 16, 9, 28, 15])
    for p in reds:
        s = p['serie']
        ws.append([p['sigla'], p['uf'], p['programa'], p['area'], p['nota'],
                   '/'.join(str(s[y]) for y in BASE), '/'.join(str(s[y]) for y in REC),
                   p['var_pct'], p['motivo'], p['cd']])
        rr = ws.max_row
        for i in range(1, 11):
            c = ws.cell(rr, i); c.border = BORDA
            c.alignment = LFT if i in (3, 4, 6, 7, 9) else CEN
        _fmt_variacao(ws.cell(rr, 8), p['var_pct'])
    return ws


def trajetoria_nacional_artefato():
    DADOS = os.path.join(AQUI, '..', 'docs', 'dados')
    sp = {y: [] for y in ANOS}
    for f in glob.glob(os.path.join(DADOS, 'area-*.json')):
        d = json.load(open(f, encoding='utf-8'))
        for r in d['data']:
            if (r.get('modalidade') or '').upper().startswith('PROFI'): continue
            for ano, vals in r.get('prod_ano', {}).items():
                mg = (vals + [0, 0, 0, 0])[2]
                if mg is not None: sp[int(ano)].append(mg)
    return {y: (round(median(v), 2) if v else None) for y, v in sp.items()}


def main():
    t0 = time.time()
    A = json.load(open(os.path.join(CACHE, 'art_total_ano.json'), encoding='utf-8'))
    art, names, ies = A['art'], A['names'], A['ies']
    roster = json.load(open(os.path.join(CACHE, 'roster_ano.json'), encoding='utf-8'))
    print('carregando metadados de programas_*.csv…')
    meta = carregar_meta_programas()
    progs = montar(art, names, ies, roster, meta)
    rets = sorted(p['ret'] for p in progs
                  if p['ret'] is not None and p['classe'] != 'ESTRUTURAL')
    nac = trajetoria_nacional_artefato()

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    aba_distribuicao(wb, progs, rets, nac)
    aba_programas(wb, progs)
    aba_resumo_ies(wb, progs)
    aba_lista_vermelha(wb, progs)
    # page setup
    from openpyxl.worksheet.properties import PageSetupProperties
    for ws in wb.worksheets:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    out = os.path.join(SAIDA, f'nacional_quedas{SUF}.xlsx')
    wb.save(out)

    from collections import Counter
    cont = Counter(p['classe'] for p in progs)
    print(f"Programas {MLABEL}: {len(progs)} | IES: {len(set(p['sigla'] for p in progs))} | " +
          ' '.join(f'{k}={cont.get(k,0)}' for k in ('VERDE', 'AMARELO', 'VERMELHO', 'ESTRUTURAL')))
    print(f"Salvo: {out}  ({time.time()-t0:.0f}s)")


if __name__ == '__main__':
    main()
