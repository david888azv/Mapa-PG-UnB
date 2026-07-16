#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PIP — Plano de Incremento de Produtividade (UnB)
================================================
Estuda quanto cada programa de pós-graduação ACADÊMICO da UnB hoje nota 3 ou 4
precisa aumentar sua produção de artigos por pesquisador para alcançar a média
nacional dos programas da nota imediatamente superior (3->4, 4->5), na MESMA
área de avaliação CAPES.

Base de dados: artefatos já preparados pelo app mapa-pg-multi
  ../docs/dados/area-*.json  (1 registro por programa x quadriênio)

Decisões metodológicas (acordadas com o usuário em 2026-06-17):
  1. NOTA VIGENTE: a nota de cada programa é a do registro do quadriênio
     2021-2024 (grade atual). A PRODUÇÃO, porém, é medida no último quadriênio
     COMPLETO (2017-2020), pois 2021-2024 ainda está com coleta incompleta.
  2. SÓ PERMANENTES: a produtividade por pesquisador usa apenas docentes
     permanentes (campos n_perm / ma_perm). ma_perm = artigos em periódico por
     docente permanente por ANO (média do quadriênio) — ATENÇÃO: o campo `ma_perm`
     do `area-*.json` NÃO é isso; ele soma os 15 subtipos de produção intelectual.
     `carregar()` o substitui pela taxa de artigos de verdade (subtipo 25, via
     prod_sub) e preserva o original em `ma_perm_prod_total`. Ver o comentário lá.
  3. SÓ ACADÊMICOS: programas profissionais ficam fora da análise e das médias
     de referência (modalidade != PROFISSIONAL).

Saída: saida/relatorio_pip_unb.xlsx
"""
import json
import glob
import os
from collections import defaultdict
from statistics import mean, median

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))
DADOS = os.path.join(AQUI, '..', 'docs', 'dados')
SAIDA = os.path.join(AQUI, 'saida')
os.makedirs(SAIDA, exist_ok=True)

QUAD_PROD = '2017-2020'   # quadriênio da produção (último completo)
QUAD_NOTA = '2021-2024'   # quadriênio da nota vigente


def is_academico(modalidade):
    """Profissional é explicitamente rotulado; tudo o mais (inclui '') é acadêmico."""
    return not (modalidade or '').upper().startswith('PROFI')


def carregar():
    """cd -> {quad: registro}, mais info de área por cd."""
    progs = defaultdict(dict)       # cd -> quad -> reg
    area_de = {}                    # cd -> (area, slug)
    areas = {}                      # slug -> nome area
    for f in sorted(glob.glob(os.path.join(DADOS, 'area-*.json'))):
        d = json.load(open(f, encoding='utf-8'))
        meta = d['metadata']
        areas[meta['slug']] = meta['area']
        for r in d['data']:
            # ── CORREÇÃO ma_perm (mesma de pip-2-refatorado-citescore/build_pip.py) ──
            # O campo `ma_perm` do `area-*.json` NÃO é uma taxa de artigos, apesar do
            # nome e da documentação. `gerar_dados_completos.py:340-362` calcula ma_*
            # filtrando só por ano e programa, SEM filtro de subtipo, e `prod.csv` traz
            # 15 subtipos de produção intelectual (25=Artigo em Periódico, 26=Capítulo,
            # 8=Resumo, 9=Congresso, 10=Texto em Jornal, e mais 10). Todo o eixo de
            # VOLUME do PIP vinha medindo produção total, não artigos.
            #
            # PPGDIP/UFRJ (2017-2020, 16 permanentes): 15 subtipos = 654 eventos →
            # ma_perm 10,22; só subtipo 25 = 428 → 6,69. Nacional, nota 5: 12,03 → 3,70.
            #
            # `prod_sub['25']['perm']` são os artigos, com paridade verificada contra a
            # soma dos estratos (311/311 em MEDICINA II) e presente nos 13.216 registros
            # com n_perm>0. Ambos contam EVENTOS artigo×autor — a unidade do app.
            # Original preservado em `ma_perm_prod_total`.
            ps = (r.get('prod_sub') or {}).get('25') or {}
            npm = r.get('n_perm') or 0
            if npm and 'perm' in ps:
                r['ma_perm_prod_total'] = r.get('ma_perm')
                r['ma_perm'] = round(ps['perm'] / npm / 4.0, 2)
            progs[r['cd']][r['quad']] = r
            area_de[r['cd']] = (meta['area'], meta['slug'])
    return progs, area_de, areas


def media_referencia(progs, area_de):
    """
    Para cada (area, nota) calcula, entre os programas ACADÊMICOS cuja NOTA
    VIGENTE (2021-2024) é essa nota, as estatísticas de ma_perm (produção
    2017-2020, art/permanente/ano): MÍNIMO (piso), MÉDIA simples, MÉDIA
    PONDERADA pelos permanentes e n.
    Retorna dict[(slug, nota)] -> {'min':, 'media':, 'wmean':, 'n':}.
    Também devolve, por área, a contagem de programas acadêmicos por nota.
    """
    grupos = defaultdict(list)      # (slug, nota_vigente) -> [(ma_perm, n_perm)]
    cont_area = defaultdict(lambda: defaultdict(int))  # slug -> nota -> n
    for cd, quads in progs.items():
        nota_reg = quads.get(QUAD_NOTA)
        prod_reg = quads.get(QUAD_PROD)
        if not nota_reg:
            continue
        if not is_academico(nota_reg.get('modalidade')):
            continue
        nota = nota_reg.get('nota')
        slug = area_de[cd][1]
        cont_area[slug][nota] += 1
        if prod_reg and prod_reg.get('n_perm', 0) and prod_reg.get('ma_perm') is not None:
            grupos[(slug, nota)].append((prod_reg['ma_perm'], prod_reg['n_perm']))
    ref = {}
    for chave, vals in grupos.items():
        mas = [m for m, _ in vals]
        peso = sum(n for _, n in vals)
        ref[chave] = {'min': round(min(mas), 2), 'media': round(mean(mas), 2),
                      'wmean': round(sum(m * n for m, n in vals) / peso, 2) if peso else None,
                      'n': len(vals)}
    return ref, cont_area


def stats_por_nota(progs):
    """
    Agrega a produção (art/permanente/ano, 2017-2020) por NOTA VIGENTE,
    nacionalmente (Brasil) e só UnB — atravessando TODAS as áreas CAPES.
    Só programas ACADÊMICOS com n_perm>0 e ma_perm definido.

    Para cada nota devolve, em 'nac' e 'unb':
      n        — nº de programas
      perm     — total de docentes permanentes (peso)
      pond     — produção PONDERADA por pesquisador/ano = Σ(ma_perm·n_perm)/Σn_perm
                 (= total de artigos/ano ÷ total de permanentes; programas maiores
                 pesam mais — é a média ponderada pedida).
      media    — média simples das taxas por programa
      mediana  — mediana das taxas por programa
    Retorna dict[nota] -> {'nac': {...}|None, 'unb': {...}|None}.
    """
    nat = defaultdict(list)         # nota -> [(ma_perm, n_perm, if_perm)]
    unb = defaultdict(list)
    for cd, quads in progs.items():
        nota_reg = quads.get(QUAD_NOTA)
        prod_reg = quads.get(QUAD_PROD)
        if not nota_reg or not is_academico(nota_reg.get('modalidade')):
            continue
        if not prod_reg or not prod_reg.get('n_perm') or prod_reg.get('ma_perm') is None:
            continue
        par = (prod_reg['ma_perm'], prod_reg['n_perm'], prod_reg.get('if_perm') or [0, 0, 0])
        nota = nota_reg.get('nota')
        nat[nota].append(par)
        if nota_reg.get('is_unb'):
            unb[nota].append(par)

    def agg(rows):
        if not rows:
            return None
        mas = [m for m, _, _ in rows]
        peso = sum(n for _, n, _ in rows)
        # perfil de impacto agregado (soma dos artigos dos permanentes por faixa)
        ift = [sum(r[2][b] for r in rows) for b in range(3)]   # [baixo, médio, alto]
        tot_if = sum(ift)
        ifp = [round(v / tot_if * 100, 1) for v in ift] if tot_if else [0, 0, 0]
        return {
            'n': len(rows),
            'perm': peso,
            'pond': round(sum(m * n for m, n, _ in rows) / peso, 2) if peso else None,
            'media': round(mean(mas), 2),
            'mediana': round(median(mas), 2),
            'if': ift,            # contagens [baixo, médio, alto]
            'if_pct': ifp,        # percentuais [baixo, médio, alto]
        }

    return {nota: {'nac': agg(nat.get(nota, [])), 'unb': agg(unb.get(nota, []))}
            for nota in (3, 4, 5, 6, 7)}


def programas_nota_unb(progs, area_de, nota):
    """Programas UnB ACADÊMICOS de uma dada NOTA VIGENTE, com produção 2017-2020
    (art/permanente/ano). Para comparar com a referência NACIONAL da MESMA nota."""
    out = []
    for cd, quads in progs.items():
        nr = quads.get(QUAD_NOTA); pr = quads.get(QUAD_PROD)
        if not nr or not nr.get('is_unb') or not is_academico(nr.get('modalidade')):
            continue
        if nr.get('nota') != nota:
            continue
        if not pr or not pr.get('n_perm') or pr.get('ma_perm') is None:
            continue
        out.append({'programa': nr.get('programa'), 'area': area_de[cd][0],
                    'n_perm': pr['n_perm'], 'ma': pr['ma_perm'],
                    'situacao': nr.get('situacao')})
    return sorted(out, key=lambda r: r['ma'])


def perfil_impacto(progs, area_de):
    """
    Perfil de impacto (OpenAlex) dos programas acadêmicos por (area, nota
    vigente), somando if_perm = [baixo, médio, alto] da produção 2017-2020 dos
    permanentes. Bandas: baixo IF<2.2 | médio 2.2-8.0 | alto >8.0.
    Retorna dict[(slug, nota)] -> [lo, mi, hi].
    """
    perf = defaultdict(lambda: [0, 0, 0])
    for cd, quads in progs.items():
        nota_reg = quads.get(QUAD_NOTA)
        prod_reg = quads.get(QUAD_PROD)
        if not nota_reg or not is_academico(nota_reg.get('modalidade')):
            continue
        if not prod_reg:
            continue
        lo, mi, hi = prod_reg.get('if_perm', [0, 0, 0])
        slug = area_de[cd][1]
        nota = nota_reg.get('nota')
        p = perf[(slug, nota)]
        p[0] += lo; p[1] += mi; p[2] += hi
    return perf


def _incremento(ma_atual, alvo_ref):
    """Retorna (incremento_num_ou_None, 'SIM'_se_ja_atingiu). Tudo art/perm/ano."""
    if ma_atual is None or alvo_ref is None:
        return None, ''
    if ma_atual >= alvo_ref:
        return 0.0, 'SIM'
    return round(alvo_ref - ma_atual, 2), ''


def analisar(progs, area_de, ref):
    """
    Linhas do relatório, para programas UnB acad. nota 3/4. Para cada um, dois
    alvos nacionais na MESMA área, em art/permanente/ano:
      - Nota 4 = MÍNIMO de ma_perm entre programas nota 4 (piso para ser nota 4).
      - Nota 5 = MÉDIA de ma_perm entre programas nota 5.
    Programa que já atinge um alvo recebe 'SIM' naquela coluna.
    """
    linhas = []
    for cd, quads in progs.items():
        nota_reg = quads.get(QUAD_NOTA)
        if not nota_reg or not nota_reg.get('is_unb'):
            continue
        if not is_academico(nota_reg.get('modalidade')):
            continue
        nota = nota_reg.get('nota')
        if nota not in (3, 4):
            continue
        area, slug = area_de[cd]

        # baseline de produção: 2017-2020; fallback 2021-2024 (com flag)
        prod_reg = quads.get(QUAD_PROD)
        baseline = QUAD_PROD
        if not prod_reg or prod_reg.get('n_perm', 0) == 0:
            prod_reg = nota_reg
            baseline = QUAD_NOTA + ' (fallback)'
        n_perm = prod_reg.get('n_perm') or 0
        ma_atual = prod_reg.get('ma_perm')
        if_perm = prod_reg.get('if_perm') or [0, 0, 0]   # [baixo, médio, alto] (2017-2020)

        # roster atual (2021-2024) para o quantitativo de docentes/bolsa
        n_doc = nota_reg.get('n_doc') or 0
        n_pq = nota_reg.get('n_pq') or 0
        n_spq = nota_reg.get('n_spq') or 0

        r4 = ref.get((slug, 4)); r5 = ref.get((slug, 5))
        alvo4 = r4['min'] if r4 else None        # piso nota 4 = mínimo nacional
        alvo5 = r5['media'] if r5 else None       # alvo nota 5 = média nacional
        n4 = r4['n'] if r4 else 0
        n5 = r5['n'] if r5 else 0
        incr4, sim4 = _incremento(ma_atual, alvo4)
        incr5, sim5 = _incremento(ma_atual, alvo5)

        linhas.append({
            'area': area, 'slug': slug, 'programa': nota_reg.get('programa'),
            'nota': nota,
            'n_doc': n_doc, 'n_pq': n_pq, 'n_spq': n_spq,
            'n_perm': n_perm, 'ma_atual': ma_atual,
            'alvo4': alvo4, 'incr4': incr4, 'sim4': sim4, 'n4': n4,
            'alvo5': alvo5, 'incr5': incr5, 'sim5': sim5, 'n5': n5,
            'baseline': baseline, 'cd': cd,
            'situacao': nota_reg.get('situacao'),
            'if_perm': if_perm,
        })
    linhas.sort(key=lambda x: (x['nota'], x['area'], x['programa']))
    return linhas

# ----------------------------- Excel ---------------------------------------

AZUL = '2C3E50'; AZUL2 = '34495E'; ROSA = 'E91E63'
H = Font(bold=True, color='FFFFFF', size=11)
HFILL = PatternFill('solid', fgColor=AZUL)
SUBFILL = PatternFill('solid', fgColor='D6EAF8')
VERDE = PatternFill('solid', fgColor='D5F5E3')
THIN = Side(style='thin', color='BBBBBB')
BORDA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


RODAPE_PERIODO = [
    'NOTA METODOLÓGICA — escolha do período de referência:',
    '• A NOTA de cada programa é a vigente (registro do quadriênio 2021-2024). '
    'A PRODUÇÃO, porém, é medida no quadriênio 2017-2020 — último período com '
    'coleta CAPES completa e confiável.',
    '• Motivo: a coleta do quadriênio 2021-2024 ainda está em andamento na '
    'Plataforma Sucupira. Apenas ~1/3 dos artigos foi lançado (a média nacional '
    'cai de ~10 para ~3 artigos/docente/ano de 2020 para 2021, sem queda real de '
    'produtividade — é subnotificação).',
    '• Além disso, a Avaliação Quadrienal 2021-2024 ainda não foi realizada pela '
    'CAPES (a de 2017-2020 só saiu em 2022; a de 2021-2024 deve sair por volta de '
    '2026-2027). As notas exibidas para 2021-2024 são o RESULTADO da avaliação de '
    '2017-2020, carregado para frente.',
    '• Logo, mede-se a produtividade no período fechado que de fato sustentou a '
    'nota vigente. Quando a CAPES consolidar 2021-2024, basta reprocessar com o '
    'novo período.',
]
ITAL = Font(italic=True, size=9, color='555555')
ITAL_B = Font(italic=True, bold=True, size=9, color=AZUL)


def rodape_periodo(ws, ncols):
    """Adiciona a nota de rodapé sobre o período, abaixo dos dados."""
    ws.append([])
    for i, texto in enumerate(RODAPE_PERIODO):
        ws.append([texto])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1)
        c.font = ITAL_B if i == 0 else ITAL
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[r].height = 14 if i == 0 else 30


def _cab(ws, headers, larguras):
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        c = ws.cell(row=ws.max_row, column=i)
        c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = larguras[i-1]
    ws.freeze_panes = 'A2'


def aba_quantitativos(wb, linhas, cont_area):
    ws = wb.create_sheet('Quantitativos UnB')
    ws.append(['QUANTITATIVOS — UnB (notas vigentes, quadriênio 2021-2024)'])
    ws.cell(1, 1).font = Font(bold=True, size=13, color=AZUL)
    ws.append([])
    # resumo por nota
    res = defaultdict(lambda: [0, 0, 0, 0])  # nota -> [n_prog, n_doc, n_pq, n_spq]
    for l in linhas:
        r = res[l['nota']]
        r[0] += 1; r[1] += l['n_doc']; r[2] += l['n_pq']; r[3] += l['n_spq']
    ws.append(['Nota', 'Nº programas', 'Total docentes', 'Com bolsa PQ', 'Sem bolsa PQ'])
    for i in range(1, 6):
        c = ws.cell(ws.max_row, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
    for nota in (3, 4):
        r = res[nota]
        ws.append([f'Nota {nota}', r[0], r[1], r[2], r[3]])
    tot = [res[3][i] + res[4][i] for i in range(4)]
    ws.append(['Total 3+4', tot[0], tot[1], tot[2], tot[3]])
    for rr in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=5):
        for c in rr:
            c.border = BORDA
            if c.column > 1: c.alignment = CENTER
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    for w, col in zip([14, 14, 16, 14, 14], 'ABCDE'):
        ws.column_dimensions[col].width = w
    rodape_periodo(ws, 5)
    return ws


def aba_incremento(wb, linhas):
    ws = wb.create_sheet('Incremento de produção')
    ws.append(['Tudo em ARTIGOS POR PESQUISADOR (permanente) POR ANO. '
               'Alvo nota 4 = mínimo nacional dos programas nota 4 na área (piso). '
               'Alvo nota 5 = média nacional dos programas nota 5 na área. '
               'SIM = já atingiu o alvo.'])
    ws.cell(1, 1).font = Font(bold=True, size=10, color=AZUL)
    ws.append([])
    headers = ['Área CAPES', 'Programa', 'Nota atual', 'Docentes perm. (2017-20)',
               'Produção atual (art/pesq/ano)',
               'Alvo NOTA 4 — mínimo nacional (art/pesq/ano)',
               'Incremento p/ NOTA 4 (art/pesq/ano)',
               'Alvo NOTA 5 — média nacional (art/pesq/ano)',
               'Incremento p/ NOTA 5 (art/pesq/ano)',
               'Baseline produção']
    larg = [25, 33, 9, 12, 14, 16, 15, 16, 15, 16]
    ws.append(headers)
    hr = ws.max_row
    for i in range(1, len(headers) + 1):
        c = ws.cell(hr, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = larg[i-1]
    ws.freeze_panes = f'A{hr+1}'

    for l in linhas:
        cel4 = 'SIM' if l['sim4'] == 'SIM' else l['incr4']
        cel5 = 'SIM' if l['sim5'] == 'SIM' else l['incr5']
        ws.append([l['area'], l['programa'], l['nota'], l['n_perm'], l['ma_atual'],
                   l['alvo4'], cel4, l['alvo5'], cel5, l['baseline']])
        row = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(row, i); c.border = BORDA
            c.alignment = LEFT if i in (1, 2) else CENTER
        if l['sim4'] == 'SIM':
            ws.cell(row, 7).fill = VERDE; ws.cell(row, 7).font = Font(bold=True, color='1E8449')
        if l['sim5'] == 'SIM':
            ws.cell(row, 9).fill = VERDE; ws.cell(row, 9).font = Font(bold=True, color='1E8449')
    rodape_periodo(ws, len(headers))
    return ws


def aba_referencias(wb, ref, areas, cont_area):
    ws = wb.create_sheet('Referências por área')
    _cab(ws, ['Área CAPES', 'Nota', 'Nº prog. acadêmicos (Brasil)',
              'Mínimo nacional art/pesq/ano (2017-20)',
              'Média nacional art/pesq/ano (2017-20)'],
         [40, 8, 16, 18, 18])
    for slug in sorted(areas):
        for nota in (3, 4, 5):
            n_prog = cont_area[slug].get(nota, 0)
            est = ref.get((slug, nota))
            if n_prog == 0 and est is None:
                continue
            ws.append([areas[slug], nota, n_prog,
                       est['min'] if est else None, est['media'] if est else None])
            for i in range(1, 6):
                c = ws.cell(ws.max_row, i); c.border = BORDA
                c.alignment = LEFT if i == 1 else CENTER
    return ws


def aba_faixas(wb, linhas, perf):
    """Distribui o incremento até a PRÓXIMA nota (art/pesquisador/ano) pelas 3
    faixas de impacto OpenAlex, segundo o perfil de impacto dos programas da
    nota-alvo na área. nota 3 -> alvo 4; nota 4 -> alvo 5."""
    ws = wb.create_sheet('Artigos por faixa de impacto')
    ws.append(['Incremento até a PRÓXIMA nota, em ARTIGOS POR PESQUISADOR POR ANO, '
               'distribuído por faixa de fator de impacto OpenAlex — '
               'baixo (IF<2,2) · médio (2,2–8,0) · alto (>8,0).'])
    ws.cell(1, 1).font = Font(bold=True, size=10, color=AZUL)
    ws.append([])
    headers = ['Área CAPES', 'Programa', 'Nota atual', 'Nota alvo (próxima)',
               'Incremento total (art/pesq/ano)',
               'Perfil nota-alvo: % baixo', '% médio', '% alto',
               'Baixo (art/pesq/ano)', 'Médio (art/pesq/ano)', 'Alto (art/pesq/ano)']
    larg = [25, 33, 9, 11, 15, 11, 10, 10, 13, 13, 13]
    ws.append(headers)
    hdr_row = ws.max_row
    for i in range(1, len(headers) + 1):
        c = ws.cell(hdr_row, i); c.font = H; c.fill = HFILL
        c.alignment = CENTER; c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = larg[i-1]
    ws.freeze_panes = f'A{hdr_row+1}'

    for l in linhas:
        alvo = l['nota'] + 1                       # 3->4, 4->5
        total = l['incr4'] if l['nota'] == 3 else l['incr5']
        lo, mi, hi = perf.get((l['slug'], alvo), [0, 0, 0])
        tot_perf = lo + mi + hi
        if total is None or tot_perf == 0:
            slo = smi = shi = None
            a_lo = a_mi = a_hi = (0 if total == 0 else None)
        else:
            slo, smi, shi = lo/tot_perf, mi/tot_perf, hi/tot_perf
            a_lo, a_mi, a_hi = round(total*slo, 1), round(total*smi, 1), round(total*shi, 1)
        ws.append([l['area'], l['programa'], l['nota'], alvo, total,
                   round(slo*100, 1) if slo is not None else None,
                   round(smi*100, 1) if smi is not None else None,
                   round(shi*100, 1) if shi is not None else None,
                   a_lo, a_mi, a_hi])
        row = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(row, i); c.border = BORDA
            c.alignment = LEFT if i in (1, 2) else CENTER
    rodape_periodo(ws, len(headers))
    return ws


NOTA_PROJ = [
    'PROJEÇÃO POR NOTA — leitura e ressalva:',
    '• "Produção ponderada" = artigos em periódico por docente PERMANENTE por ANO '
    '(2017-2020), agregando TODAS as áreas CAPES e ponderando cada programa pelo '
    'seu nº de permanentes (= total de artigos/ano ÷ total de permanentes). É a '
    'média de produção PEDIDA, ponderada pelos programas de cada nota.',
    '• A projeção pega a produção ponderada ATUAL dos programas UnB de cada nota e '
    'mede quanto falta (art/pesq/ano) para alcançar a MÉDIA ou a MEDIANA NACIONAL '
    'dos programas da nota imediatamente superior (3→4, 4→5, 5→6, 6→7).',
    '• RESSALVA IMPORTANTE: a produção por pesquisador SATURA a partir da nota 4 '
    '(Brasil: ~10,9 na nota 4 e ~11–12 nas notas 5/6/7). Acima da nota 4 os '
    'programas se distinguem mais pelo IMPACTO/qualidade do que pelo VOLUME por '
    'pesquisador — então o incremento de quantidade aqui é condição necessária, '
    'não suficiente, para subir de 5 em diante. Ver a aba de faixas de impacto.',
]


def aba_projecao(wb, stats):
    """Produção ponderada por nota (Brasil × UnB) + projeção de incremento da
    UnB até a média/mediana nacional da próxima nota."""
    ws = wb.create_sheet('Projeção por nota')
    ws.append(['PRODUÇÃO PONDERADA POR NOTA E PROJEÇÃO DE INCREMENTO (UnB)'])
    ws.cell(1, 1).font = Font(bold=True, size=13, color=AZUL)
    ws.append([])

    # ---- Tabela 1: produção por nota, Brasil x UnB -----------------------
    h1 = ['Nota', 'Prog. Brasil', 'Perm. Brasil',
          'Ponderada Brasil (art/pesq/ano)', 'Média Brasil', 'Mediana Brasil',
          'Prog. UnB', 'Perm. UnB', 'Ponderada UnB (art/pesq/ano)']
    larg1 = [7, 11, 12, 16, 12, 13, 9, 10, 16]
    ws.append(h1)
    hr = ws.max_row
    for i in range(1, len(h1) + 1):
        c = ws.cell(hr, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = larg1[i-1]
    for nota in (7, 6, 5, 4, 3):
        nac = stats[nota]['nac']; unb = stats[nota]['unb']
        ws.append([
            nota,
            nac['n'] if nac else 0, nac['perm'] if nac else 0,
            nac['pond'] if nac else None, nac['media'] if nac else None,
            nac['mediana'] if nac else None,
            unb['n'] if unb else 0, unb['perm'] if unb else 0,
            unb['pond'] if unb else None,
        ])
        for i in range(1, len(h1) + 1):
            c = ws.cell(ws.max_row, i); c.border = BORDA; c.alignment = CENTER

    # ---- Tabela 2: projeção de incremento UnB ----------------------------
    ws.append([])
    ws.append(['PROJEÇÃO — produção ponderada ATUAL da UnB → alvo nacional da PRÓXIMA nota'])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color=ROSA)
    h2 = ['De → Para', 'Prog. UnB', 'Ponderada UnB atual (art/pesq/ano)',
          'Alvo MÉDIA nacional (próx. nota)', 'Incremento p/ MÉDIA (art/pesq/ano)',
          'Alvo MEDIANA nacional (próx. nota)', 'Incremento p/ MEDIANA (art/pesq/ano)',
          'Incremento % (sobre a média)']
    larg2 = [11, 9, 16, 16, 16, 16, 16, 14]
    ws.append(h2)
    hr2 = ws.max_row
    for i in range(1, len(h2) + 1):
        c = ws.cell(hr2, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
        atual = ws.column_dimensions[get_column_letter(i)].width or 0
        ws.column_dimensions[get_column_letter(i)].width = max(atual, larg2[i-1])

    for g in (3, 4, 5, 6):
        unb = stats[g]['unb']; alvo = stats[g + 1]['nac']
        if not unb or not alvo:
            continue
        base = unb['pond']
        a_med = alvo['media']; a_mdn = alvo['mediana']
        inc_med = round(max(0.0, a_med - base), 2)
        inc_mdn = round(max(0.0, a_mdn - base), 2)
        pct = round(inc_med / base * 100, 1) if base else None
        ws.append([f'{g} → {g+1}', unb['n'], base, a_med, inc_med, a_mdn, inc_mdn, pct])
        row = ws.max_row
        for i in range(1, len(h2) + 1):
            c = ws.cell(row, i); c.border = BORDA; c.alignment = CENTER
        # destaca quando a UnB já está no/acima do alvo (incremento 0)
        if inc_med == 0:
            ws.cell(row, 5).fill = VERDE; ws.cell(row, 5).font = Font(bold=True, color='1E8449')
        if inc_mdn == 0:
            ws.cell(row, 7).fill = VERDE; ws.cell(row, 7).font = Font(bold=True, color='1E8449')

    # nota de rodapé específica
    ws.append([])
    for i, texto in enumerate(NOTA_PROJ):
        ws.append([texto])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(h2))
        c = ws.cell(r, 1)
        c.font = ITAL_B if i == 0 else ITAL
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[r].height = 14 if i == 0 else 42
    return ws


def aba_detalhe_transicao(wb, linhas, stats):
    """Detalhe POR PROGRAMA UnB das transições 3→4 e 4→5: produção atual vs a
    mediana/média nacional da nota-alvo (mesmos alvos da aba 'Projeção por nota'),
    incremento por pesquisador/ano e em artigos/ano absolutos. Ordena do mais
    perto ao mais longe. Programas em desativação e novos (sem produção 2017-2020)
    ficam fora do ranking, listados à parte."""
    ws = wb.create_sheet('Detalhe 3→4 e 4→5')
    ws.append(['DETALHE POR PROGRAMA UnB — incremento de produção até a PRÓXIMA nota'])
    ws.cell(1, 1).font = Font(bold=True, size=13, color=AZUL)
    ws.append(['Produção em artigos por permanente/ano (2017-2020). Alvo = MEDIANA e MÉDIA '
               'nacionais dos programas da nota-alvo (mesmos valores da aba "Projeção por nota"). '
               '"+ artigos/ano" = incremento × nº de permanentes. Ordenado do mais perto ao mais longe.'])
    ws.cell(2, 1).font = ITAL

    headers = ['Área CAPES', 'Programa', 'Nota', 'Perm. (2017-20)',
               'Produção atual (art/pesq/ano)', 'Alvo MEDIANA (próx.)',
               'Incr. p/ mediana (art/pesq/ano)', '+ artigos/ano (mediana)',
               'Alvo MÉDIA (próx.)', 'Incr. p/ média (art/pesq/ano)',
               '+ artigos/ano (média)', 'Situação / obs.']
    larg = [24, 32, 6, 11, 13, 12, 14, 13, 12, 14, 13, 24]

    for nota in (3, 4):
        alvo = stats[nota + 1]['nac']
        med, mean = alvo['mediana'], alvo['media']
        ws.append([])
        ws.append([f'NOTA {nota} → {nota+1}   (alvo nacional: mediana {med} · média {mean} art/pesq/ano)'])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color=ROSA)
        ws.append(headers)
        hr = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(hr, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
            col = get_column_letter(i)
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, larg[i-1])

        grupo = [l for l in linhas if l['nota'] == nota]
        ativos = [l for l in grupo if l['situacao'] == 'EM FUNCIONAMENTO'
                  and 'fallback' not in l['baseline'] and l['ma_atual'] is not None]
        especiais = [l for l in grupo if l not in ativos]
        ativos.sort(key=lambda l: max(0.0, med - l['ma_atual']))

        tot_med = tot_mean = tot_perm = 0
        for l in ativos:
            ma = l['ma_atual']; n = l['n_perm'] or 0
            im = round(max(0.0, med - ma), 2); imn = round(max(0.0, mean - ma), 2)
            am = round(im * n); amn = round(imn * n)
            tot_med += am; tot_mean += amn; tot_perm += n
            status = '✓ já ≥ mediana' if im == 0 else ''
            ws.append([l['area'], l['programa'], nota, n, ma, med, im, am, mean, imn, amn, status])
            row = ws.max_row
            for i in range(1, len(headers) + 1):
                c = ws.cell(row, i); c.border = BORDA
                c.alignment = LEFT if i in (1, 2, 12) else CENTER
            if im == 0:
                ws.cell(row, 7).fill = VERDE; ws.cell(row, 8).fill = VERDE

        ws.append([f'SUBTOTAL — {len(ativos)} programas ativos', '', '', tot_perm, '', '',
                   '', tot_med, '', '', tot_mean, 'artigos/ano a mais (coorte)'])
        sr = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(sr, i); c.border = BORDA; c.font = Font(bold=True)
            c.alignment = LEFT if i in (1, 12) else CENTER

        for l in especiais:
            if l['situacao'] != 'EM FUNCIONAMENTO':
                obs = 'EM DESATIVAÇÃO — fora do plano'
            else:
                obs = 'programa novo — sem produção 2017-2020'
            ma = l['ma_atual'] if ('fallback' not in l['baseline'] and l['ma_atual'] is not None) else '—'
            ws.append([l['area'], l['programa'], nota, l['n_perm'] or 0, ma, '', '', '', '', '', '', obs])
            row = ws.max_row
            for i in range(1, len(headers) + 1):
                c = ws.cell(row, i); c.border = BORDA
                c.alignment = LEFT if i in (1, 2, 12) else CENTER
                c.font = Font(italic=True, color='777777')

    rodape_periodo(ws, len(headers))
    return ws


def aba_piso(wb, linhas, ref):
    """Incremento de cada programa UnB nota 3/4 até o PISO da nota-alvo NA PRÓPRIA
    área CAPES = a MENOR produção (art/perm/ano) entre os programas da nota-alvo
    na mesma área (regra de comparação da CAPES). Mostra também a média ponderada
    (pelos permanentes) da nota-alvo na área. Programas em desativação / novos
    ficam fora do ranking."""
    ws = wb.create_sheet('Piso da área (3→4, 4→5)')
    ws.append(['INCREMENTO ATÉ O PISO DA NOTA-ALVO NA PRÓPRIA ÁREA CAPES'])
    ws.cell(1, 1).font = Font(bold=True, size=13, color=AZUL)
    ws.append(['Produção em art/permanente/ano (2017-2020). Para cada programa UnB o ALVO é o '
               'PISO = MENOR produção entre os programas da nota-alvo NA MESMA área CAPES. '
               'Inclui também a MÉDIA PONDERADA (pelos permanentes) da nota-alvo na área. '
               '"+ art/ano" = incremento × nº de permanentes. Ordenado do mais perto ao mais longe.'])
    ws.cell(2, 1).font = ITAL

    headers = ['Área CAPES', 'Programa', 'Nota', 'Perm. (2017-20)', 'Produção atual',
               'PISO nota-alvo na área', 'Incr. p/ piso', '+ art/ano (piso)',
               'Méd. ponderada nota-alvo', 'Incr. p/ méd. pond.', '+ art/ano (méd. pond.)',
               'nº prog. nota-alvo na área']
    larg = [24, 32, 6, 11, 12, 13, 12, 13, 14, 13, 14, 13]

    for nota in (3, 4):
        alvo_nota = nota + 1
        ws.append([])
        ws.append([f'NOTA {nota} → {alvo_nota}   (alvo = piso/menor produção dos programas nota {alvo_nota} na mesma área)'])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color=ROSA)
        ws.append(headers)
        hr = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(hr, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
            col = get_column_letter(i)
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, larg[i-1])

        grupo = [l for l in linhas if l['nota'] == nota]
        ativos = [l for l in grupo if l['situacao'] == 'EM FUNCIONAMENTO'
                  and 'fallback' not in l['baseline'] and l['ma_atual'] is not None]
        especiais = [l for l in grupo if l not in ativos]

        def sortkey(l):
            r = ref.get((l['slug'], alvo_nota))
            if not r or r['min'] <= 0:          # sem ref ou piso=0 (não comparável) ao fim
                return (1, 0.0)
            return (0, max(0.0, r['min'] - l['ma_atual']))
        ativos.sort(key=sortkey)

        tot_piso = tot_wm = tot_perm = n_ref = 0
        for l in ativos:
            ma = l['ma_atual']; n = l['n_perm'] or 0
            r = ref.get((l['slug'], alvo_nota))
            if not r:
                ws.append([l['area'], l['programa'], nota, n, ma, '—', '—', '—', '—', '—', '—',
                           f'sem programa nota {alvo_nota} na área'])
                row = ws.max_row
                for i in range(1, len(headers) + 1):
                    c = ws.cell(row, i); c.border = BORDA
                    c.alignment = LEFT if i in (1, 2, 12) else CENTER
                    c.font = Font(italic=True, color='777777')
                continue
            piso = r['min']; wm = r['wmean']
            piso_ok = piso > 0                  # piso=0 não é referência útil
            wm_ok = wm is not None and wm > 0
            if piso_ok:
                ip = round(max(0.0, piso - ma), 2); ap = round(ip * n); tot_piso += ap
                piso_c, ip_c, ap_c = piso, ip, ap
            else:
                ip = None; piso_c = piso; ip_c = ap_c = '—'
            if wm_ok:
                iw = round(max(0.0, wm - ma), 2); aw = round(iw * n); tot_wm += aw
                wm_c, iw_c, aw_c = wm, iw, aw
            else:
                iw = None; wm_c = wm if wm is not None else '—'; iw_c = aw_c = '—'
            tot_perm += n; n_ref += 1
            obs = '' if piso_ok else 'piso = 0 (sem referência útil)'
            ws.append([l['area'], l['programa'], nota, n, ma, piso_c, ip_c, ap_c, wm_c, iw_c, aw_c,
                       obs or r['n']])
            row = ws.max_row
            for i in range(1, len(headers) + 1):
                c = ws.cell(row, i); c.border = BORDA
                c.alignment = LEFT if i in (1, 2) else CENTER
            if piso_ok and ip == 0:
                ws.cell(row, 7).fill = VERDE; ws.cell(row, 8).fill = VERDE

        ws.append([f'SUBTOTAL — {n_ref} prog. com referência na área', '', '', tot_perm, '', '',
                   '', tot_piso, '', '', tot_wm, 'artigos/ano a mais (coorte)'])
        sr = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(sr, i); c.border = BORDA; c.font = Font(bold=True)
            c.alignment = LEFT if i in (1, 12) else CENTER

        for l in especiais:
            obs = ('EM DESATIVAÇÃO — fora do plano' if l['situacao'] != 'EM FUNCIONAMENTO'
                   else 'programa novo — sem produção 2017-2020')
            ma = l['ma_atual'] if ('fallback' not in l['baseline'] and l['ma_atual'] is not None) else '—'
            ws.append([l['area'], l['programa'], nota, l['n_perm'] or 0, ma, '', '', '', '', '', '', obs])
            row = ws.max_row
            for i in range(1, len(headers) + 1):
                c = ws.cell(row, i); c.border = BORDA
                c.alignment = LEFT if i in (1, 2, 12) else CENTER
                c.font = Font(italic=True, color='777777')

    rodape_periodo(ws, len(headers))
    return ws


def _quatro_alvos(l, ref, stats):
    """Para uma linha (programa UnB nota 3/4) devolve os 4 alvos da nota seguinte
    e o incremento (art/pesq/ano) até cada um: piso e média ponderada da PRÓPRIA
    área; mediana e média NACIONAIS. None onde não há referência."""
    nota = l['nota']; alvo = nota + 1; ma = l['ma_atual']
    r = ref.get((l['slug'], alvo))
    nac = stats[alvo]['nac']
    # piso só vale como referência se for > 0: um piso = 0 (algum programa da
    # nota-alvo sem produção) é trivialmente superado por qualquer programa, logo
    # não serve de meta — vira None (sem comparação).
    piso = r['min'] if r else None
    if piso is not None and piso <= 0:
        piso = None
    alvos = {
        'piso':  piso,
        'wmean': r['wmean'] if r else None,
        'mediana': nac['mediana'] if nac else None,
        'media':   nac['media'] if nac else None,
    }
    incr = {k: (round(max(0.0, v - ma), 2) if v is not None else None) for k, v in alvos.items()}
    return alvos, incr


def aba_comparativo(wb, linhas, ref, stats):
    """Consolida as QUATRO definições de alvo (piso da área, média ponderada da
    área, mediana nacional, média nacional) numa única tabela: por programa, o
    incremento (art/pesq/ano) até cada alvo; e o resumo das metas ABSOLUTAS
    (artigos/ano a mais) por coorte. Programas em desativação/novos ficam fora."""
    ws = wb.create_sheet('Comparativo de alvos')
    ws.append(['COMPARATIVO DAS 4 DEFINIÇÕES DE ALVO — incremento por pesquisador/ano'])
    ws.cell(1, 1).font = Font(bold=True, size=13, color=AZUL)
    ws.append(['Produção em art/permanente/ano (2017-2020). Cada coluna de incremento mostra quanto '
               'falta (art/pesq/ano) para o programa alcançar aquele alvo da NOTA SEGUINTE. '
               'Piso e Média ponderada são da PRÓPRIA área CAPES; Mediana e Média são NACIONAIS. '
               'Rigor crescente: piso < (média pond. / mediana) < média.'])
    ws.cell(2, 1).font = ITAL

    # ── Resumo de coorte: artigos/ano a mais por alvo ──────────────────────
    ws.append([])
    ws.append(['RESUMO — meta da COORTE em artigos/ano a mais (soma por programa, piso em zero)'])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color=ROSA)
    resumo_hdr = ['Transição', 'Programas', '→ Piso da área', '→ Média ponderada da área',
                  '→ Mediana nacional', '→ Média nacional']
    ws.append(resumo_hdr)
    rhr = ws.max_row
    for i in range(1, len(resumo_hdr) + 1):
        c = ws.cell(rhr, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA

    cohort = {}     # nota -> dict de totais absolutos
    for nota in (3, 4):
        ativos = [l for l in linhas if l['nota'] == nota and l['situacao'] == 'EM FUNCIONAMENTO'
                  and 'fallback' not in l['baseline'] and l['ma_atual'] is not None]
        tot = {'piso': 0, 'wmean': 0, 'mediana': 0, 'media': 0}
        for l in ativos:
            _, incr = _quatro_alvos(l, ref, stats)
            n = l['n_perm'] or 0
            for k in tot:
                if incr[k] is not None:
                    tot[k] += round(incr[k] * n)
        cohort[nota] = (len(ativos), tot)
        ws.append([f'{nota} → {nota+1}', len(ativos),
                   tot['piso'], tot['wmean'], tot['mediana'], tot['media']])
        for i in range(1, len(resumo_hdr) + 1):
            c = ws.cell(ws.max_row, i); c.border = BORDA; c.alignment = CENTER
    for col, w in zip('ABCDEF', [12, 11, 14, 22, 16, 14]):
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, w)

    # ── Detalhe por programa ───────────────────────────────────────────────
    headers = ['Área CAPES', 'Programa', 'Nota', 'Perm. (2017-20)', 'Produção atual',
               'Incr → PISO área', 'Incr → MÉD.POND. área', 'Incr → MEDIANA nac.', 'Incr → MÉDIA nac.']
    larg = [24, 32, 6, 11, 12, 13, 15, 14, 13]
    for nota in (3, 4):
        ws.append([])
        nac = stats[nota + 1]['nac']
        ws.append([f'NOTA {nota} → {nota+1}   (alvos nacionais: mediana {nac["mediana"]} · média {nac["media"]} art/pesq/ano)'])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color=ROSA)
        ws.append(headers)
        hr = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(hr, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
            col = get_column_letter(i)
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, larg[i-1])

        grupo = [l for l in linhas if l['nota'] == nota]
        ativos = [l for l in grupo if l['situacao'] == 'EM FUNCIONAMENTO'
                  and 'fallback' not in l['baseline'] and l['ma_atual'] is not None]
        especiais = [l for l in grupo if l not in ativos]
        ativos.sort(key=lambda l: (_quatro_alvos(l, ref, stats)[1]['mediana'] or 0))

        for l in ativos:
            _, incr = _quatro_alvos(l, ref, stats)
            cell = lambda k: ('—' if incr[k] is None else incr[k])
            ws.append([l['area'], l['programa'], nota, l['n_perm'] or 0, l['ma_atual'],
                       cell('piso'), cell('wmean'), cell('mediana'), cell('media')])
            row = ws.max_row
            for i in range(1, len(headers) + 1):
                c = ws.cell(row, i); c.border = BORDA
                c.alignment = LEFT if i in (1, 2) else CENTER
            # verde nas colunas cujo incremento é 0 (já atingiu)
            for col_i, k in ((6, 'piso'), (7, 'wmean'), (8, 'mediana'), (9, 'media')):
                if incr[k] == 0:
                    ws.cell(row, col_i).fill = VERDE

        for l in especiais:
            obs = ('EM DESATIVAÇÃO — fora do plano' if l['situacao'] != 'EM FUNCIONAMENTO'
                   else 'programa novo — sem produção 2017-2020')
            ws.append([l['area'], l['programa'], nota, l['n_perm'] or 0, '—', '', '', '', obs])
            row = ws.max_row
            for i in range(1, len(headers) + 1):
                c = ws.cell(row, i); c.border = BORDA
                c.alignment = LEFT if i in (1, 2) else CENTER
                c.font = Font(italic=True, color='777777')

    rodape_periodo(ws, len(headers))
    return ws


def aba_nota5(wb, progs, area_de, stats):
    """Programas nota 5 da UnB vs a referência NACIONAL da nota 5 (ponderada,
    média e mediana). Mostra quantos estão ABAIXO da mediana/média — argumento de
    que, na nota 5, a produção já não discrimina (o foco passa a ser infraestrutura)."""
    nac = stats[5]['nac']
    med, mean, pond = nac['mediana'], nac['media'], nac['pond']
    progs5 = programas_nota_unb(progs, area_de, 5)

    ws = wb.create_sheet('Nota 5 vs nacional')
    ws.append(['PROGRAMAS NOTA 5 DA UnB — produção vs referência NACIONAL da nota 5'])
    ws.cell(1, 1).font = Font(bold=True, size=13, color=AZUL)
    n_med = sum(1 for p in progs5 if p['ma'] < med)
    n_mean = sum(1 for p in progs5 if p['ma'] < mean)
    ws.append([f'Referência nacional nota 5 (art/perm/ano, 2017-2020): ponderada {pond} · '
               f'média {mean} · mediana {med}. Dos {len(progs5)} programas nota 5 da UnB, '
               f'{n_med} estão ABAIXO da mediana e {n_mean} abaixo da média — ou seja, a maioria '
               f'alcançou a nota 5 com produção abaixo da referência: o volume já não é o que '
               f'distingue a nota. O passo seguinte (5→6→7) depende de IMPACTO e INFRAESTRUTURA.'])
    ws.cell(2, 1).font = ITAL

    headers = ['Programa', 'Área CAPES', 'Perm. (2017-20)', 'Produção (art/pesq/ano)',
               'Mediana nac.', 'Falta p/ mediana', 'Média nac.', 'Falta p/ média', 'Situação']
    larg = [34, 30, 11, 14, 11, 13, 11, 13, 26]
    ws.append(headers)
    hr = ws.max_row
    for i in range(1, len(headers) + 1):
        c = ws.cell(hr, i); c.font = H; c.fill = HFILL; c.alignment = CENTER; c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = larg[i-1]
    ws.freeze_panes = f'A{hr+1}'

    for p in progs5:
        ma = p['ma']
        fm = round(med - ma, 2) if ma < med else 0.0
        fmn = round(mean - ma, 2) if ma < mean else 0.0
        if ma >= mean:
            stt = '≥ mediana e média'
        elif ma >= med:
            stt = 'abaixo só da média'
        else:
            stt = 'abaixo da mediana e da média'
        ws.append([p['programa'], p['area'], p['n_perm'], ma, med,
                   fm if fm > 0 else '—', mean, fmn if fmn > 0 else '—', stt])
        row = ws.max_row
        for i in range(1, len(headers) + 1):
            c = ws.cell(row, i); c.border = BORDA
            c.alignment = LEFT if i in (1, 2, 9) else CENTER
        if ma >= med:
            ws.cell(row, 6).fill = VERDE
        if ma >= mean:
            ws.cell(row, 8).fill = VERDE

    ws.append([f'RESUMO — {len(progs5)} programas: {len(progs5)-n_med} ≥ mediana · {n_med} abaixo · '
               f'{len(progs5)-n_mean} ≥ média · {n_mean} abaixo', '', '', '', '', '', '', '', ''])
    sr = ws.max_row
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
    ws.cell(sr, 1).font = Font(bold=True)
    return ws


def main():
    progs, area_de, areas = carregar()
    ref, cont_area = media_referencia(progs, area_de)
    perf = perfil_impacto(progs, area_de)
    stats = stats_por_nota(progs)
    linhas = analisar(progs, area_de, ref)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    aba_quantitativos(wb, linhas, cont_area)
    aba_incremento(wb, linhas)
    aba_projecao(wb, stats)
    aba_detalhe_transicao(wb, linhas, stats)
    aba_piso(wb, linhas, ref)
    aba_comparativo(wb, linhas, ref, stats)
    aba_nota5(wb, progs, area_de, stats)
    aba_faixas(wb, linhas, perf)
    aba_referencias(wb, ref, areas, cont_area)
    out = os.path.join(SAIDA, 'relatorio_pip_unb.xlsx')
    wb.save(out)

    # resumo no terminal
    n3 = [l for l in linhas if l['nota'] == 3]
    n4 = [l for l in linhas if l['nota'] == 4]
    print(f'Programas UnB acadêmicos nota 3: {len(n3)} | nota 4: {len(n4)}')
    ja4 = [l for l in linhas if l['sim4'] == 'SIM']
    ja5 = [l for l in linhas if l['sim5'] == 'SIM']
    print(f'Já atingem o PISO de nota 4 (mínimo nacional): {len(ja4)}')
    print(f'Já atingem a MÉDIA de nota 5 (média nacional): {len(ja5)}')
    print('\nProjeção por nota (produção ponderada art/pesq/ano, 2017-2020):')
    print('nota | pondBR | médiaBR | medianaBR | pondUnB')
    for nota in (7, 6, 5, 4, 3):
        nac = stats[nota]['nac']; u = stats[nota]['unb']
        print(f'  {nota}  | {nac["pond"]:6.2f} | {nac["media"]:6.2f}  | '
              f'{nac["mediana"]:6.2f}    | {u["pond"] if u else float("nan"):6.2f}')
    print('Incremento UnB (ponderada atual -> média/mediana nacional da próxima nota):')
    for g in (3, 4, 5, 6):
        u = stats[g]['unb']; alvo = stats[g + 1]['nac']
        if not u or not alvo:
            continue
        base = u['pond']
        print(f'  {g}->{g+1}: UnB {base:.2f} | +{max(0, alvo["media"]-base):.2f} p/ média '
              f'({alvo["media"]:.2f}) | +{max(0, alvo["mediana"]-base):.2f} p/ mediana '
              f'({alvo["mediana"]:.2f})')
    print(f'Excel salvo em: {out}')


if __name__ == '__main__':
    main()
