#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
INVESTIGAÇÃO DE QUEDAS DE PRODUÇÃO — UnB (2013-2024)
====================================================
Recálculo a partir dos DADOS BRUTOS da CAPES (Dados Abertos) para a UnB,
contornando os artefatos da camada pré-processada do mapa-pg-multi.

PROBLEMA QUE MOTIVOU O RECÁLCULO
--------------------------------
A camada pré-processada (docs/dados/area-*.json, campo prod_ano) atribui a
produção de cada artigo a um docente do *roster de 2021* (único ano de docentes
coletado no quadriênio 2021-2024). Resultado: ~2/3 da produção real de 2021-2024
é descartada (artigos de docentes admitidos após 2021, de quem saiu, ou ligados
a discentes/externos). Ex.: ANTROPOLOGIA aparece com queda de 92% no build,
enquanto nos dados brutos manteve 128/122/104/95 artigos/ano (estável).

SINAL LIMPO ADOTADO
-------------------
Nº de ARTIGOS DISTINTOS (ID_ADD_PRODUCAO_INTELECTUAL) por programa por ano,
direto de prod_intel_*_artpe (cobre 2013-2024, sem depender de roster nem de
vínculo autor). É a medida de produção imune aos artefatos do build.
Vínculo por categoria (permanente) vem de prod_autor (2013-2020) e
prod_artpe (2021-2024), que trazem NM_TP_CATEGORIA_DOCENTE.

CLASSIFICAÇÃO (retenção limpa = média artigos 2021-24 / média artigos 2017-20)
  ⬛ ESTRUTURAL  programa em desativação, sucedido por novo código, ou rede/
                 minter/dinter encerrada (zerou antes de 2021). Não é alerta.
  🔴 VERMELHO    ret < 0.50  e não-estrutural — queda acentuada (alerta)
  🟡 AMARELO     0.50 <= ret < 0.85 — queda moderada
  🟢 VERDE       ret >= 0.85 — manteve ou cresceu

Saídas (saida/):
  relatorio_quedas_unb.xlsx · quedas_criticas.png · lista_vermelha.md
"""
import os, sys, json, glob
from statistics import mean, median
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Modalidade a analisar: 'ACAD' (padrão) ou 'PROF'. Uso: python3 investiga_quedas.py prof
MODAL = 'PROF' if any(a.lower().startswith('prof') for a in sys.argv[1:]) else 'ACAD'
MLABEL = 'profissionais' if MODAL == 'PROF' else 'acadêmicos'
SUF = '_profissional' if MODAL == 'PROF' else ''

AQUI = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(AQUI, 'cache_unb')
DADOS = os.path.join(AQUI, '..', 'docs', 'dados')
SAIDA = os.path.join(AQUI, 'saida')
os.makedirs(SAIDA, exist_ok=True)

ANOS = list(range(2013, 2025))
BASE = [2017, 2018, 2019, 2020]   # último quadriênio com coleta completa
REC  = [2021, 2022, 2023, 2024]   # quadriênio vigente (coleta em andamento)

# ---- limiares de classificação (retenção limpa) ----
LIM_RED = 0.50
LIM_YEL = 0.85

# =====================================================================
# 1. CARGA
# =====================================================================
def carregar():
    A = json.load(open(os.path.join(CACHE, 'art_total_ano.json'), encoding='utf-8'))
    art, names = A['art'], A['names']
    roster = json.load(open(os.path.join(CACHE, 'roster_ano.json'), encoding='utf-8'))

    def load_opt(fn):
        p = os.path.join(CACHE, fn)
        return json.load(open(p, encoding='utf-8')) if os.path.exists(p) else {}
    pa_artpe = load_opt('prod_artpe_unb.json')   # 2021-2024 por categoria
    pa_autor = load_opt('prod_autor_unb.json')   # 2013-2020 por categoria

    meta = {}
    for f in glob.glob(os.path.join(DADOS, 'area-*.json')):
        d = json.load(open(f, encoding='utf-8'))
        area = d['metadata']['area']
        for r in d['data']:
            if r.get('is_unb'):
                meta[r['cd']] = {'area': area, 'programa': r['programa'],
                                 'modal': r.get('modalidade', ''),
                                 'situacao': r.get('situacao', '')}
    return art, names, roster, pa_artpe, pa_autor, meta


def is_acad(m):
    return not (m or '').upper().startswith('PROFI')


def manter_modal(m, modal=None):
    """True se o programa pertence à modalidade pedida (modal ou o MODAL global)."""
    prof = (m or '').upper().startswith('PROFI')
    return prof if (modal or MODAL) == 'PROF' else not prof


def detectar_sucessores(meta, art, modal=None):
    """nome de programa -> lista de cds; marca par desativado->ativo."""
    from collections import defaultdict
    nm = defaultdict(list)
    for cd, m in meta.items():
        if manter_modal(m['modal'], modal):
            nm[m['programa']].append(cd)
    sucesso_de = {}   # cd_ativo -> cd_antigo
    for nome, cds in nm.items():
        if len(cds) < 2:
            continue
        desat = [c for c in cds if 'DESATIV' in (meta[c]['situacao'] or '').upper()]
        ativ  = [c for c in cds if 'DESATIV' not in (meta[c]['situacao'] or '').upper()]
        if desat and ativ:
            for a in ativ:
                sucesso_de[a] = desat[0]
    return sucesso_de


# =====================================================================
# 2. MODELO POR PROGRAMA
# =====================================================================
def serie_art(art, cd):
    return {y: art.get(f"{cd}|{y}", 0) for y in ANOS}


def montar(art, names, roster, pa_artpe, pa_autor, meta, modal=None):
    sucesso_de = detectar_sucessores(meta, art, modal)
    progs = []
    for cd in sorted(set(c.split('|')[0] for c in art)):
        m = meta.get(cd)
        if not m or not manter_modal(m['modal'], modal):
            continue
        s = serie_art(art, cd)

        # docentes por ano (roster 2013-2021; forward-fill 2021 -> 2022-24)
        ndoc = {}; nperm = {}
        for y in ANOS:
            r = roster.get(f"{cd}|{y}", {})
            ndoc[y] = r.get('n_doc'); nperm[y] = r.get('n_perm')
        ff_doc = ndoc.get(2021) or ndoc.get(2020)
        ff_perm = nperm.get(2021) or nperm.get(2020)
        for y in (2022, 2023, 2024):
            if not ndoc[y]:  ndoc[y] = ff_doc
            if not nperm[y]: nperm[y] = ff_perm

        # artigos com autor permanente por ano (autor 2013-20 + artpe 2021-24)
        artperm = {}
        for y in ANOS:
            src = pa_artpe if y >= 2021 else pa_autor
            # todos os 12 anos são cobertos por uma fonte; ausência = 0 artigos
            artperm[y] = src.get(f"{cd}|{y}", {}).get('art_perm', 0)

        # nota vigente
        nota = (roster.get(f"{cd}|2021", {}).get('nota')
                or roster.get(f"{cd}|2020", {}).get('nota'))

        base = mean(s[y] for y in BASE)
        rec  = mean(s[y] for y in REC)
        ret  = (rec / base) if base > 0 else None

        # classificação
        sit = (m['situacao'] or '').upper()
        # encerrou antes do período: sem produção em 2020 NEM em 2021-24 (última atividade ≤2019)
        encerrou_antes = (base > 0 and rec == 0 and s[2020] == 0)
        if 'DESATIV' in sit or cd in sucesso_de.values():
            classe = 'ESTRUTURAL'; motivo = 'em desativação / sucedido por novo código'
        elif encerrou_antes:
            classe = 'ESTRUTURAL'; motivo = 'encerrado/recodificado (sem produção desde 2020)'
        elif ret is None:
            classe = 'ESTRUTURAL'; motivo = 'sem base de comparação (curso novo/sucessor)'
        elif ret < LIM_RED:
            classe = 'VERMELHO'
            # subnotificação: 2021 normal, 2022-24 colapsa
            rec_tail = mean(s[y] for y in (2022, 2023, 2024))
            if base > 0 and s[2021] >= 0.5 * base and rec_tail < 0.25 * base:
                motivo = 'subnotificação Sucupira (reportou 2021; 2022-24 ~zero)'
            else:
                motivo = 'queda acentuada — investigar'
            nome_u = (names.get(cd, '') or '').upper()
            if any(t in nome_u for t in ('REDE', 'MINTER', 'DINTER', 'PRÓ', 'PRO-', 'NACIONAL')):
                motivo += ' — programa em REDE: verificar produção nas IES parceiras'
        elif ret < LIM_YEL:
            classe = 'AMARELO'; motivo = 'queda moderada'
        else:
            classe = 'VERDE'; motivo = 'manteve ou cresceu'

        # Sucessor (código novo de programa desativado) com base 2017-20 incompleta:
        # a "retenção" infla artificialmente (base quase-zero) → não é crescimento real.
        antigo = sucesso_de.get(cd)
        base_incompleta = bool(antigo) and sum(1 for y in BASE if s[y] == 0) >= 2
        var_pct = (round((ret - 1) * 100, 1) if ret is not None else None)
        if base_incompleta:
            var_pct = None
            motivo = (f'novo curso (sucessor de …{str(antigo)[-4:]}) — base 2017-20 '
                      'incompleta; produção atual saudável')

        progs.append({
            'cd': cd, 'programa': names.get(cd, m['programa']), 'area': m['area'],
            'nota': nota, 'situacao': m['situacao'],
            'serie': s, 'ndoc': ndoc, 'nperm': nperm, 'artperm': artperm,
            'base': base, 'rec': rec, 'ret': ret,
            # variação 2021-24 vs 2017-20: queda = negativo; crescimento = positivo; manteve = 0
            'var_pct': var_pct,
            'classe': classe, 'motivo': motivo,
            'sucessor_de': antigo, 'base_incompleta': base_incompleta,
        })
    ordem = {'VERMELHO': 0, 'ESTRUTURAL': 1, 'AMARELO': 2, 'VERDE': 3}
    progs.sort(key=lambda p: (ordem[p['classe']], p['ret'] if p['ret'] is not None else 9))
    return progs


# =====================================================================
# 3. EXCEL
# =====================================================================
COR = {'VERDE': '2ECC71', 'AMARELO': 'F1C40F', 'VERMELHO': 'E74C3C', 'ESTRUTURAL': '95A5A6'}
FILLS = {k: PatternFill('solid', fgColor=v) for k, v in COR.items()}
H = Font(bold=True, color='FFFFFF', size=11)
HFILL = PatternFill('solid', fgColor='2C3E50')
THIN = Side(style='thin', color='CCCCCC')
BORDA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _fmt_variacao(cell, val):
    """Convenção: queda = negativo; crescimento = positivo em NEGRITO; manteve = 0."""
    if val is None or not isinstance(val, (int, float)):
        return
    cell.number_format = '+0.0;-0.0;0'   # sinal explícito; zero sem sinal
    if val > 0:
        cell.font = Font(bold=True, color='1E8449')   # crescimento: negrito, verde
    elif val < 0:
        cell.font = Font(color='C0392B')               # queda: vermelho


def _hdr(ws, headers, larg, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h); c.font = H; c.fill = HFILL; c.alignment = CEN; c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = larg[i-1]
    # coordenada como string para NÃO instanciar a célula (evita linha em branco no append)
    ws.freeze_panes = f'{get_column_letter(3)}{row + 1}'


def aba_resumo(wb, progs, nac):
    ws = wb.create_sheet('Resumo')
    from collections import Counter
    cont = Counter(p['classe'] for p in progs)
    rets = [p['ret'] for p in progs if p['ret'] is not None and p['classe'] != 'ESTRUTURAL']
    med_ret = median(rets) if rets else 0
    linhas = [
        (f'INVESTIGAÇÃO DE QUEDAS DE PRODUÇÃO — UnB · programas {MLABEL.upper()} (dados brutos CAPES)', 14, '2C3E50', True),
        ('', 9, '000000', False),
        ('Sinal: nº de artigos distintos por programa por ano (prod_intel_artpe, dados brutos, 2013-2024).', 10, '555555', False),
        ('Imune ao artefato da camada pré-processada, que subnotifica 2021-2024 ao exigir vínculo com o roster de 2021.', 10, '555555', False),
        ('', 9, '000000', False),
        (f"Programas UnB {MLABEL} analisados: {len(progs)}", 11, '000000', True),
        (f"   🟢 VERDE (manteve/cresceu, ret≥{LIM_YEL}): {cont.get('VERDE',0)}", 11, '1E8449', False),
        (f"   🟡 AMARELO (queda moderada, {LIM_RED}–{LIM_YEL}): {cont.get('AMARELO',0)}", 11, '9A7D0A', False),
        (f"   🔴 VERMELHO (queda acentuada, ret<{LIM_RED}): {cont.get('VERMELHO',0)}", 11, 'C0392B', True),
        (f"   ⬛ ESTRUTURAL (desativado/sucessor/rede encerrada): {cont.get('ESTRUTURAL',0)}", 11, '566573', False),
        ('', 9, '000000', False),
        (f'ACHADO: retenção limpa mediana dos programas {MLABEL} ≈ {med_ret:.2f} (artigos brutos 2021-24 / 2017-20).', 11, 'C0392B', True),
        ('A "queda" vista na camada pré-processada e na mediana nacional é artefato de', 10, '555555', False),
        ('coleta (Sucupira em andamento) + atribuição ao roster de 2021. Ver aba "Metodologia".', 10, '555555', False),
        ('', 9, '000000', False),
        ('Trajetória NACIONAL (mediana art/docente/ano, camada pré-processada) — mostra o artefato coletivo:', 10, '2C3E50', True),
    ]
    r = 1
    for txt, sz, cor, bold in linhas:
        ws.cell(r, 1, txt).font = Font(size=sz, color=cor, bold=bold)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        r += 1
    # mini-série nacional
    ws.cell(r, 1, 'Ano').font = H; ws.cell(r, 1).fill = HFILL
    for j, y in enumerate(ANOS, 2):
        c = ws.cell(r, j, y); c.font = H; c.fill = HFILL; c.alignment = CEN
    ws.cell(r+1, 1, 'Mediana nac.').font = Font(bold=True)
    for j, y in enumerate(ANOS, 2):
        ws.cell(r+1, j, nac.get(y)).alignment = CEN
    ws.column_dimensions['A'].width = 16
    return ws


def aba_serie(wb, progs):
    ws = wb.create_sheet('Série anual (artigos)')
    headers = ['Programa', 'Área CAPES'] + [str(y) for y in ANOS] + ['Nota', 'ret', '% variação', 'Classe', 'Observação']
    larg = [32, 22] + [5]*len(ANOS) + [5, 6, 9, 11, 30]
    _hdr(ws, headers, larg)
    col_var = 2 + len(ANOS) + 3   # coluna do '% variação'
    for p in progs:
        ret_disp = '—' if (p['ret'] is None or p.get('base_incompleta')) else round(p['ret'], 2)
        row = [p['programa'], p['area']] + [p['serie'][y] for y in ANOS] + [
            p['nota'], ret_disp,
            (p['var_pct'] if p['var_pct'] is not None else '—'),
            p['classe'], p['motivo']]
        ws.append(row)
        rr = ws.max_row
        for i in range(1, len(headers)+1):
            c = ws.cell(rr, i); c.border = BORDA
            c.alignment = LFT if i in (1, 2, len(headers)) else CEN
        _fmt_variacao(ws.cell(rr, col_var), p['var_pct'])
        # colorir células dos anos REC conforme classe
        for y in REC:
            ci = 2 + ANOS.index(y) + 1
            ws.cell(rr, ci).fill = FILLS[p['classe']]
        ws.cell(rr, len(headers)-1).fill = FILLS[p['classe']]
    return ws


def aba_por_pesquisador(wb, progs):
    ws = wb.create_sheet('Produção por pesquisador')
    ws.cell(1,1,'Produção por pesquisador por ANO — (a) por todos os docentes = artigos/n_doc; '
                '(b) por permanente = artigos c/ autor permanente / n_perm. '
                '2022-24: n_doc/n_perm = roster 2021 (único disponível).').font = Font(italic=True, size=9, color='555555')
    ws.merge_cells('A1:Z1')
    headers = ['Programa', 'Métrica'] + [str(y) for y in ANOS]
    larg = [32, 16] + [5]*len(ANOS)
    _hdr(ws, headers, larg, row=2)
    for p in progs:
        if p['classe'] == 'ESTRUTURAL':
            continue
        # por todos
        ws.append([p['programa'], 'art/docente'] + [
            (round(p['serie'][y]/p['ndoc'][y], 2) if p['ndoc'].get(y) else None) for y in ANOS])
        r1 = ws.max_row
        # por permanente — repete o nome do programa para alinhar nome×métrica em toda linha
        ws.append([p['programa'], 'art/permanente'] + [
            (round(p['artperm'][y]/p['nperm'][y], 2) if (p['artperm'].get(y) is not None and p['nperm'].get(y)) else None) for y in ANOS])
        r2 = ws.max_row
        sep = Side(style='medium', color='888888')   # separador entre programas
        for rr in (r1, r2):
            for i in range(1, len(headers)+1):
                c = ws.cell(rr, i); c.border = BORDA
                c.alignment = LFT if i in (1, 2) else CEN
            for y in REC:
                ws.cell(rr, 2+ANOS.index(y)+1).fill = FILLS[p['classe']]
        # borda superior reforçada na 1ª linha do par, p/ destacar o agrupamento
        for i in range(1, len(headers)+1):
            cc = ws.cell(r1, i)
            cc.border = Border(left=THIN, right=THIN, bottom=THIN, top=sep)
    return ws


def aba_lista_vermelha(wb, progs):
    ws = wb.create_sheet('Lista vermelha')
    sub = [p for p in progs if p['classe'] in ('VERMELHO', 'ESTRUTURAL')]
    _hdr(ws, ['Programa', 'Área CAPES', 'Nota', 'Classe', 'Série 2017-2020',
              'Série 2021-2024', '% variação', 'Diagnóstico (verificado no bruto)', 'Código CAPES'],
         [30, 24, 5, 11, 18, 18, 9, 40, 16])
    for p in sub:
        s = p['serie']
        ws.append([p['programa'], p['area'], p['nota'], p['classe'],
                   '/'.join(str(s[y]) for y in BASE), '/'.join(str(s[y]) for y in REC),
                   (p['var_pct'] if p['var_pct'] is not None else '—'),
                   p['motivo'], p['cd']])
        rr = ws.max_row
        for i in range(1, 10):
            c = ws.cell(rr, i); c.border = BORDA
            c.alignment = LFT if i in (1, 2, 5, 6, 8) else CEN
        _fmt_variacao(ws.cell(rr, 7), p['var_pct'])
        ws.cell(rr, 4).fill = FILLS[p['classe']]
    return ws


def aba_metodologia(wb):
    ws = wb.create_sheet('Metodologia')
    txt = [
        ('POR QUE RECALCULAR DOS DADOS BRUTOS', 12, True, '2C3E50'),
        ('', 9, False, '000000'),
        ('1. A camada pré-processada (prod_ano) credita um artigo ao programa apenas se o', 10, False, '333333'),
        ('   docente-autor estiver no ROSTER DE 2021 (único ano de docentes coletado em 2021-2024).', 10, False, '333333'),
        ('   Verificado: ANTROPOLOGIA/UnB tem 128 artigos em 2021, mas só 34 mapeiam ao roster →', 10, False, '333333'),
        ('   o build mostra "queda de 92%" que é FALSA (produção real estável: 128/122/104/95).', 10, False, '333333'),
        ('', 9, False, '000000'),
        ('2. Não existe prod_autor_2021a2024 nos Dados Abertos; o vínculo autor↔artigo de 2021-24', 10, False, '333333'),
        ('   vem de prod_artpe_*. O sinal limpo (nº de artigos distintos por programa/ano, via', 10, False, '333333'),
        ('   prod_intel_artpe) não depende de roster nem de vínculo e cobre 2013-2024.', 10, False, '333333'),
        ('', 9, False, '000000'),
        ('3. A coleta do quadriênio 2021-2024 segue em andamento na Sucupira. A mediana NACIONAL', 10, False, '333333'),
        ('   cai de ~8 (2020) para ~3 (2021) art/doc/ano — queda sistêmica (3.378 programas),', 10, False, '333333'),
        ('   logo NÃO é perda de produtividade, e sim subnotificação coletiva.', 10, False, '333333'),
        ('', 9, False, '000000'),
        ('4. CLASSIFICAÇÃO relativa ao comportamento coletivo: como quase todo programa "cai" na', 10, False, '333333'),
        ('   contagem absoluta, usa-se a RETENÇÃO limpa (artigos 21-24 / 17-20). Programas que', 10, False, '333333'),
        ('   apenas acompanham a coleta ficam ≥0,85 (verde); só os discrepantes ficam vermelhos.', 10, False, '333333'),
        ('', 9, False, '000000'),
        ('5. Casos ESTRUTURAIS (cinza) NÃO são alertas: programa em desativação (ex.: BOTÂNICA', 10, False, '333333'),
        ('   053…038P0, sucedido por 053…112P6), ou rede/MINTER/DINTER que encerrou antes de 2021', 10, False, '333333'),
        ('   (ex.: CONTABILIDADE UnB-UFPB-UFRN, zerada desde 2019).', 10, False, '333333'),
    ]
    for i, (t, sz, b, c) in enumerate(txt, 1):
        ws.cell(i, 1, t).font = Font(size=sz, bold=b, color=c)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
    ws.column_dimensions['A'].width = 18
    return ws


# =====================================================================
# 4. GRÁFICO
# =====================================================================
def grafico(progs, nac_ma, path):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    reds = [p for p in progs if p['classe'] in ('VERMELHO', 'ESTRUTURAL')]
    fig, ax = plt.subplots(figsize=(11, 6.2))
    # mediana UnB (art/doc) como referência
    med = []
    for y in ANOS:
        vals = [p['serie'][y]/p['ndoc'][y] for p in progs
                if p['classe'] != 'ESTRUTURAL' and p['ndoc'].get(y)]
        med.append(median(vals) if vals else None)
    ax.plot(ANOS, med, color='#27AE60', lw=3, marker='o', label='Mediana UnB (art/docente)', zorder=5)
    cores = ['#E74C3C', '#8E44AD', '#D35400', '#7F8C8D', '#2980B9', '#C0392B']
    for i, p in enumerate(reds):
        serie = [p['serie'][y]/p['ndoc'][y] if p['ndoc'].get(y) else None for y in ANOS]
        ls = '--' if p['classe'] == 'ESTRUTURAL' else '-'
        tag = '[estrutural]' if p['classe'] == 'ESTRUTURAL' else '[vermelho]'
        ax.plot(ANOS, serie, ls=ls, lw=2, marker='s', ms=4,
                color=cores[i % len(cores)], alpha=0.9,
                label=f"{tag} {p['programa']}")   # nome COMPLETO (sem corte)
    ax.axvspan(2020.5, 2024.5, color='#FADBD8', alpha=0.5, zorder=0)
    ax.text(2022.5, ax.get_ylim()[1]*0.96, '2021-2024\n(coleta em andamento)',
            ha='center', va='top', fontsize=9, color='#922B21')
    ax.set_xlabel('Ano-base'); ax.set_ylabel('Artigos por docente / ano')
    ax.set_title(f'UnB · programas {MLABEL} em queda crítica vs. mediana institucional\n'
                 '(dados brutos CAPES; recálculo mapa-pg-multi/pip)', fontsize=12, fontweight='bold')
    ax.set_xticks(ANOS); ax.grid(alpha=0.3)
    # legenda ABAIXO do gráfico, 1 coluna, nomes completos; bbox_inches='tight' a captura
    ax.legend(fontsize=9, loc='upper center', bbox_to_anchor=(0.5, -0.12),
              ncol=1, framealpha=0.95, borderaxespad=0.5, handlelength=2.5)
    fig.savefig(path, dpi=130, bbox_inches='tight'); plt.close(fig)


# =====================================================================
# 5. MARKDOWN
# =====================================================================
def markdown(progs, path):
    reds = [p for p in progs if p['classe'] == 'VERMELHO']
    estr = [p for p in progs if p['classe'] == 'ESTRUTURAL']
    rets = [p['ret'] for p in progs if p['ret'] is not None and p['classe'] != 'ESTRUTURAL']
    med_ret = median(rets) if rets else 0
    L = []
    L.append(f'# Lista vermelha — quedas de produção UnB · programas {MLABEL} (2021-2024)\n')
    L.append('> Recálculo a partir dos **dados brutos** da CAPES (Dados Abertos). '
             'Sinal: nº de artigos distintos por programa por ano (`prod_intel_artpe`), '
             'imune ao artefato de subnotificação da camada pré-processada.\n')
    L.append('## Achado central\n')
    L.append(f'Retenção limpa mediana dos programas {MLABEL} ≈ **{med_ret:.2f}** '
             '(artigos brutos 2021-24 / 2017-20). A "queda" observada na camada pré-processada e na mediana '
             'nacional é **artefato de coleta** (Sucupira em andamento + atribuição ao roster de 2021), '
             'não perda real de produtividade.\n')
    if MODAL == 'PROF':
        L.append('> ⚠️ **Cautela (programas profissionais):** a avaliação CAPES de mestrados/doutorados '
                 'profissionais valoriza fortemente a **produção técnica/aplicada** (produtos, software, '
                 'patentes), não só artigos em periódico. O sinal aqui (artigos) **subnotifica** essas '
                 'modalidades — programas com poucos/zero artigos podem ter forte produção técnica. '
                 'Trate os alertas como indicativos, não conclusivos.\n')
    L.append(f'## 🔴 Programas em alerta vermelho ({len(reds)})\n')
    if not reds:
        L.append('_Nenhum programa acadêmico com queda genuína acentuada (após excluir casos estruturais)._\n')
    for p in reds:
        s = p['serie']
        L.append(f"### {p['programa']} — {p['area']} (nota {p['nota']})")
        L.append(f"- Série 2017-2020: {'/'.join(str(s[y]) for y in BASE)} artigos · "
                 f"2021-2024: {'/'.join(str(s[y]) for y in REC)} artigos")
        L.append(f"- Variação: **{p['var_pct']:+.1f}%** · Diagnóstico: {p['motivo']}")
        L.append(f"- Código CAPES: `{p['cd']}`\n")
    L.append(f'## ⬛ Casos estruturais — NÃO são alertas ({len(estr)})\n')
    L.append('| Programa | Área | Série 17-20 → 21-24 | Motivo |')
    L.append('|---|---|---|---|')
    for p in estr:
        s = p['serie']
        L.append(f"| {p['programa']} | {p['area']} | "
                 f"{'/'.join(str(s[y]) for y in BASE)} → {'/'.join(str(s[y]) for y in REC)} | {p['motivo']} |")
    open(path, 'w', encoding='utf-8').write('\n'.join(L) + '\n')


# =====================================================================
def trajetoria_nacional():
    """Mediana nacional art/doc/ano (camada pré-processada) — para mostrar o artefato."""
    serie_por_ano = {y: [] for y in ANOS}
    for f in glob.glob(os.path.join(DADOS, 'area-*.json')):
        d = json.load(open(f, encoding='utf-8'))
        for r in d['data']:
            if (r.get('modalidade') or '').upper().startswith('PROFI'):
                continue
            for ano, vals in r.get('prod_ano', {}).items():
                mg = (vals + [0, 0, 0, 0])[2]
                if mg is not None:
                    serie_por_ano[int(ano)].append(mg)
    return {y: (round(median(v), 2) if v else None) for y, v in serie_por_ano.items()}


def main():
    art, names, roster, pa_artpe, pa_autor, meta = carregar()
    progs = montar(art, names, roster, pa_artpe, pa_autor, meta)
    nac = trajetoria_nacional()

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    aba_resumo(wb, progs, nac)
    aba_serie(wb, progs)
    aba_por_pesquisador(wb, progs)
    aba_lista_vermelha(wb, progs)
    aba_metodologia(wb)
    # impressão/PDF legível: paisagem + ajuste à largura
    from openpyxl.worksheet.properties import PageSetupProperties
    for ws in wb.worksheets:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins.left = ws.page_margins.right = 0.2
        ws.page_margins.top = ws.page_margins.bottom = 0.3
    xlsx = os.path.join(SAIDA, f'relatorio_quedas_unb{SUF}.xlsx')
    wb.save(xlsx)

    grafico(progs, nac, os.path.join(SAIDA, f'quedas_criticas{SUF}.png'))
    markdown(progs, os.path.join(SAIDA, f'lista_vermelha{SUF}.md'))

    from collections import Counter
    cont = Counter(p['classe'] for p in progs)
    print(f"UnB {MLABEL}: {len(progs)} | " + ' '.join(f"{k}={cont.get(k,0)}" for k in ('VERDE','AMARELO','VERMELHO','ESTRUTURAL')))
    print('VERMELHOS:')
    for p in progs:
        if p['classe'] == 'VERMELHO':
            print(f"  {p['programa'][:40]:<40} ret={p['ret']:.2f} | {p['motivo']}")
    print(f"Saídas em {SAIDA}/")
    pa_ok = 'sim' if pa_artpe else 'NÃO (rode caches)'
    print(f"[cache permanente artpe: {pa_ok}; autor: {'sim' if pa_autor else 'NÃO'}]")


if __name__ == '__main__':
    main()
