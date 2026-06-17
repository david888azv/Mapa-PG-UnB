#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera saida/estatisticas_unb_por_nota.xlsx — os programas de pós-graduação da
UnB EM FUNCIONAMENTO, organizados por nota CAPES vigente.

Reproduz (de forma versionada) a planilha que antes era um one-off sem script.

Duas abas:
  • "Programas (detalhe)" : 1 linha por programa (92), ordenada por nota desc e,
    dentro da nota, pelo NOME do programa na ordem bruta do Python (por ponto de
    código, sensível a acento); empates de nome preservam a ordem do registry
    (sort estável → ADMINP antes de ADMIN).
  • "Resumo por nota"      : contagens/somatórios por nota via fórmulas VIVAS
    (COUNTIF/SUMIF sobre a aba de detalhe), de modo que editar o detalhe recalcula.

Fontes (sem reprocessar dados brutos):
  ../docs/registry.json          -> sufixo, nome, área, grande área, modalidade,
                                     graus e conceito (NOTA VIGENTE = conceito[-1],
                                     o último quadriênio; DESCOO/MECATR=['3','4']→4).
  ../docs/dados/area-<slug>.json -> n_doc / n_perm do quadriênio 2021-2024
                                     (roster 2021 — único ano de docentes coletado
                                     no quadriênio). Programa-rede sem fatia UnB
                                     (ex.: CONTABILIDADE UnB-UFPB-UFRN) fica vazio.
"""
import os
import json

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)                       # mapa-pg-multi/
REGISTRY = os.path.join(RAIZ, 'docs', 'registry.json')
DADOS = os.path.join(RAIZ, 'docs', 'dados')
SAIDA = os.path.join(AQUI, 'saida')
DEST = os.path.join(SAIDA, 'estatisticas_unb_por_nota.xlsx')

QUAD_VIGENTE = '2021-2024'                          # roster mais recente


def nota_vigente(conceito):
    """A NOTA vigente é o ÚLTIMO conceito da lista (cronológica por quadriênio):
    DESCOO/MECATR = ['3','4'] -> a nota atual é 4, não 3."""
    return (conceito or [''])[-1]


def rank_nota(conceito):
    """Ordena 7>6>5>4>3 e joga 'A' (programas novos, sem nota) para o fim."""
    c = nota_vigente(conceito)
    return int(c) if c.isdigit() else -1


# --- cache de roster (n_doc / n_perm) por área, lido sob demanda ------------
_cache_area = {}


def roster_2021(slug, cd_programa):
    """Retorna (n_doc, n_perm) do registro 2021-2024 do programa, ou (None,None)."""
    if slug not in _cache_area:
        with open(os.path.join(DADOS, 'area-%s.json' % slug), encoding='utf-8') as f:
            _cache_area[slug] = json.load(f)['data']
    for r in _cache_area[slug]:
        if r['cd'] == cd_programa and r['quad'] == QUAD_VIGENTE:
            return r.get('n_doc'), r.get('n_perm')
    return None, None


def montar_linhas():
    with open(REGISTRY, encoding='utf-8') as f:
        progs = json.load(f)['programas_unb']

    linhas = []
    for p in progs:
        n_doc, n_perm = roster_2021(p['slug_area'], p['cd_programa'])
        linhas.append({
            'sufixo': p['sufixo'],
            'programa': p['nome'],
            'area': p['area_capes'],
            'grande_area': p['grande_area_cnpq'],
            'modalidade': ', '.join(p['modalidade']),
            'graus': ', '.join(p['graus']),
            'nota': nota_vigente(p['conceito']),
            'n_doc': n_doc,
            'n_perm': n_perm,
        })

    # nota desc, depois nome do programa em ordem bruta do Python (por ponto de
    # código, portanto SENSÍVEL a acento — "CIENCIAS MEDICAS" vem antes de
    # "CIÊNCIAS ..." porque 'E' < 'Ê'); empate de nome mantém a ordem do registry
    # porque sorted() é estável e as chaves são aplicadas nessa ordem.
    linhas.sort(key=lambda l: l['programa'])
    linhas.sort(key=lambda l: rank_nota([l['nota']]), reverse=True)
    return linhas


def gerar():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    linhas = montar_linhas()
    n = len(linhas)                                 # 92
    fim = n + 1                                     # última linha de dados (header=1)

    AZUL = PatternFill('solid', fgColor='305496')
    BRANCO_BOLD = Font(bold=True, color='FFFFFF')
    THIN = Side(style='thin')
    BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    CENTRO = Alignment(horizontal='center')

    wb = Workbook()

    # ----------------------------------------------------------------- detalhe
    ws = wb.active
    ws.title = 'Programas (detalhe)'
    cabec = ['Sufixo', 'Programa', 'Área CAPES', 'Grande área CNPq', 'Modalidade',
             'Graus', 'Nota', 'Docentes (2021)', 'Permanentes (2021)']
    ws.append(cabec)
    for c in ws[1]:
        c.font = BRANCO_BOLD
        c.fill = AZUL
        c.alignment = CENTRO
        c.border = BOX

    for l in linhas:
        ws.append([l['sufixo'], l['programa'], l['area'], l['grande_area'],
                   l['modalidade'], l['graus'], l['nota'],
                   l['n_doc'] if l['n_doc'] is not None else None,
                   l['n_perm'] if l['n_perm'] is not None else None])

    for row in ws.iter_rows(min_row=2, max_row=fim, max_col=9):
        for c in row:
            c.border = BOX
            if c.column_letter in ('G', 'H', 'I'):
                c.alignment = CENTRO

    for col, w in zip('ABCDEFGHI', [9, 52, 46, 34, 13, 22, 6, 15, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'

    # ------------------------------------------------------------------ resumo
    rs = wb.create_sheet('Resumo por nota')
    rs['A1'] = 'Programas de pós-graduação da UnB EM FUNCIONAMENTO, por nota CAPES'
    rs['A1'].font = Font(bold=True, size=13)
    rs['A2'] = ('Nota vigente = registro 2021-2024 (registry.json, %d programas). '
                'Docentes = roster 2021.' % n)
    rs['A2'].font = Font(italic=True, color='808080', size=9)

    cab = ['Nota', 'Programas', 'Acadêmicos', 'Profissionais', 'Docentes',
           'Permanentes', 'Méd. doc/prog']
    for j, t in enumerate(cab, start=1):
        c = rs.cell(row=4, column=j, value=t)
        c.font = BRANCO_BOLD
        c.fill = AZUL
        c.alignment = CENTRO
        c.border = BOX

    # referências absolutas à aba de detalhe
    D = "'Programas (detalhe)'"
    G = "%s!$G$2:$G$%d" % (D, fim)      # Nota
    E = "%s!$E$2:$E$%d" % (D, fim)      # Modalidade
    H = "%s!$H$2:$H$%d" % (D, fim)      # Docentes
    I = "%s!$I$2:$I$%d" % (D, fim)      # Permanentes

    notas = [('7', '7'), ('6', '6'), ('5', '5'), ('4', '4'), ('3', '3'),
             ('A', 'A (novos)')]
    r = 5
    for chave, rotulo in notas:
        rs.cell(row=r, column=1, value=rotulo)
        rs.cell(row=r, column=2,
                value='=COUNTIF(%s,"%s")' % (G, chave))
        rs.cell(row=r, column=3,
                value='=COUNTIFS(%s,"%s",%s,"ACADÊMICO")' % (G, chave, E))
        rs.cell(row=r, column=4,
                value='=COUNTIFS(%s,"%s",%s,"PROFISSIONAL")' % (G, chave, E))
        rs.cell(row=r, column=5,
                value='=SUMIF(%s,"%s",%s)' % (G, chave, H))
        rs.cell(row=r, column=6,
                value='=SUMIF(%s,"%s",%s)' % (G, chave, I))
        g = rs.cell(row=r, column=7,
                    value='=IFERROR(E%d/COUNTIFS(%s,"%s",%s,"<>"),"")'
                          % (r, G, chave, H))
        g.number_format = '0.0'
        r += 1

    # TOTAL
    rs.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
    for col in 'BCDEF':
        rs['%s%d' % (col, r)] = '=SUM(%s5:%s%d)' % (col, col, r - 1)
    gt = rs.cell(row=r, column=7,
                 value='=IFERROR(E%d/COUNTIFS(%s,"<>"),"")' % (r, H))
    gt.number_format = '0.0'
    total_row = r

    # estilo das linhas do corpo do resumo: borda em tudo, células numéricas
    # centradas; a linha TOTAL ganha realce (negrito + fundo azul-claro) e seu
    # rótulo "TOTAL" na coluna A não é centralizado (como no original).
    REALCE = PatternFill('solid', fgColor='D9E1F2')
    for row in rs.iter_rows(min_row=5, max_row=total_row, max_col=7):
        for c in row:
            c.border = BOX
            if c.row == total_row:
                c.font = Font(bold=True)
                c.fill = REALCE
                if c.column != 1:
                    c.alignment = CENTRO
            else:
                c.alignment = CENTRO

    for col, w in zip('ABCDEFG', [12, 11, 12, 14, 11, 13, 14]):
        rs.column_dimensions[col].width = w

    # notas de rodapé
    notas_rodape = [
        '',
        'Notas:',
        '• "A" = 3 programas novos aprovados (APCN), ainda sem nota quadrienal — '
        'inclui a nova Botânica (sucessora da desativada) e 2 profissionais.',
        '• Nota 5 tem 1 programa sem dado de docentes: CONTABILIDADE em rede '
        'UnB-UFPB-UFRN (multi-IES, fatia UnB zerada); a média e os totais o ignoram.',
        '• Docentes = roster 2021, único ano de coleta de docentes no quadriênio '
        '2021-2024 (limitação dos datasets CAPES).',
        '• Excluídos 3 CDs em desativação que ainda surgem na métrica (Botânica '
        'antiga, Turismo prof., Tecnologias Quím/Biol).',
        '• Fórmulas vivas: este resumo recalcula a partir da aba "Programas (detalhe)".',
    ]
    CINZA = Font(italic=True, color='595959', size=9)
    linha = total_row + 1
    for txt in notas_rodape:
        if txt:
            rs.cell(row=linha, column=1, value=txt).font = CINZA
        linha += 1

    os.makedirs(SAIDA, exist_ok=True)
    wb.save(DEST)
    return n


if __name__ == '__main__':
    n = gerar()
    print('OK — %d programas -> %s' % (n, DEST))
